#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SIM - Sistemi intonazione musicale
Copyright (c) 2025 Luca Bimbi
Distribuito secondo la licenza MIT - vedi il file LICENSE per i dettagli

Nome programma: SIM - Sistemi intonazione musicale
Versione: 1.5
Autore: LUCA BIMBI
Data: 2025-08-30

Descrizione
-----------
Generatore di parametri/rapporti utili alla costruzione di tabelle di accordatura
(per esempio per cpstun in Csound). Supporta due modalità principali:

1) Temperamento equabile (ET)
   - Calcolo del rapporto della radice per un temperamento equabile generico
     specificando l'indice della radice (es. 12 per 12-TET) e l'ampiezza dell'ottava
     in cents (es. 1200), oppure una frazione razionale (es. 3/2) da convertire in cents.
   - Esempio: -et 12 1200

2) Sistema geometrico su generatore
   - Genera una progressione geometrica a partire da un generatore (es. 3/2) per un numero
     di passi desiderato, con riduzione all'ottava attiva di default (ricondotti in [1, 2)).
   - Opzione --no-reduce per mantenere i rapporti non ridotti (crescenti senza ricondurli in ottava).
   - Esempio: --geometric 3/2 12

3) Sistema di intonazione naturale (4:5:6)
   - Rapporti generati con formula ((3/2)^a) * ((5/4)^b).
   - Intervalli esplorati variando a in [-A_MAX...A_MAX] e b in [-B_MAX...B_MAX].
   - Esempio: --natural 3 3

4) Sistema Danielou
   - Formula: ((6/5)^a) * (3/2)^b * 2^c, con c usato per la riduzione in ottava.
   - Default: sottoinsieme a partire dalla tonica, tre terze minori armoniche successive,
     tre seste maggiori armoniche successive e l'asse delle quinte (a=0, b=-5..5).
   - Opzione --danielou-all per includere la griglia completa a∈[-3..3], b∈[-5..5] (fino a 53 rapporti).
   - Esponenti opzionali: usa "--danielou a,b,c" per generare un rapporto specifico.

Input e parametri principali
---------------------------
- --basenote: nota di riferimento come nome (es. "A4", "F#2") oppure frequenza in Hz (float).
- --diapason: diapason in Hz usato per convertire i nomi di nota in Hz (default: 440).
- --basekey: nota base (MIDI key) per tabella cpstun; attualmente è un parametro informativo
  lasciato per compatibilità d'uso (default: 60 = C4).
- -et INDEX Cents|Frazione: imposta l'ET; i cents possono essere interi o una frazione (es. 3/2).
- --geometric GEN STEPS: imposta il sistema geometrico (GEN può essere intero, frazione o float;
  STEPS è il numero di passi, > 0). Riduzione all'ottava attiva di default.
- --natural A_MAX B_MAX: sistema naturale 4:5:6 con ((3/2)^a)*((5/4)^b), a∈[-A_MAX...A_MAX], b∈[-B_MAX...B_MAX].
- --danielou: genera il sistema Danielou (sottoinsieme predefinito); aggiungi --danielou-all per la griglia completa (fino a 53 rapporti).
- --no-reduce: disattiva la riduzione all'ottava in modalità generazione (default: riduci).
- output_file: nome base del file di output; verrà creato/aggiornato "<output_file>.csd" con tabelle cpstun (righe 'f') di tipo GEN -2 contenenti i rapporti (ratios) generati.

Note su frazioni e cents
------------------------
- Se per --et vengono passati i cents come frazione razionale (Fraction), questa viene convertita in cents
  tramite il logaritmo naturale secondo la scala di Ellis.
- Conversioni disponibili: ratio -> cents (fraction_to_cents) e cents -> ratio approssimato (cents_to_fraction).

Esempi d'uso
------------
- Geometrico con riduzione in ottava (default):
  python3 sim.py --basekey 60 --basenote 440 --geometric 3/2 12 out

- Geometrico senza riduzione in ottava:
  python3 sim.py --basekey 60 --basenote 440 --geometric 3/2 5 --no-reduce out2

- Sistema naturale 4:5:6 (a, b in [-3..3]):
  python3 sim.py --basekey 60 --basenote 440 --natural 3 3 out_nat

- Sistema Danielou (sottoinsieme):
  python3 sim.py --basekey 60 --basenote 440 --danielou out_dan

- Sistema Danielou (griglia completa, fino a 53 rapporti):
  python3 sim.py --basekey 60 --basenote 440 --danielou-all out_dan_all

- Temperamento equabile 12-TET (ottava 1200 cents):
  python3 sim.py --basekey 60 --basenote 440 -et 12 1200 out_et

Output
------
- Stampa a video le frequenze calcolate per ciascun passo.
- Scrive/aggiorna un file "<output_file>.csd" aggiungendo una riga di score per cpstun:
  una f‑tabella di tipo GEN -2 che contiene i seguenti campi all'interno della lista dati:
  [numgrades, interval, basefreq, basekey, r1, r2, ...].
  - numgrades = numero di rapporti generati;
  - interval = intervallo di ripetizione del sistema (es. 2 per ottava, 1.5 per quinta). Se il comando
    conosce l'intervallo (ET o sistemi normalizzati all'ottava), viene scritto esplicitamente;
    in caso contrario, viene inferito come 2.0 quando tutti i rapporti sono in [1,2] (inclusivo) e 0.0
    quando non determinabile;
  - basefreq = frequenza assoluta associata a basekey (Hz);
  - basekey = nota MIDI di riferimento.
  Ogni esecuzione sullo stesso file aggiunge una nuova tabella con numero (f_max + 1). I valori r1..rn sono i
  rapporti (unitless) rispetto a basefreq.
- Esporta tabelle del sistema generato (Step, MIDI, Ratio, Hz) in "<output_file>_system.txt" e, se disponibile openpyxl,
  anche "<output_file>_system.xlsx".
- Esporta tabelle di confronto con 12TET, serie armonica e serie subarmonica (incluse colonne DeltaHz) in
  "<output_file>_compare.txt" e, se disponibile openpyxl, anche "<output_file>_compare.xlsx"
  disponendo valori vicini affiancati (Custom accanto a Harmonic/Subharmonic) e usando colori distinti
  (rosso per Custom, verde per Harmonics, giallo per Subharmonics, blu per TET), con evidenziazione quando le differenze sono molto piccole.
- Opzionale: esporta un file ".tun" (AnaMark TUN) con la mappatura assoluta di tutte le 128 note MIDI
  usando --export-tun. Le altezze dei passi generati vengono assegnate a partire da basekey,
  con frequenze assolute calcolate come basefrequency * ratio; le altre note vengono riempite
  in 12TET relativo alla stessa basefrequency.

"""

from __future__ import annotations

import argparse
import sys
import math
from fractions import Fraction
import re
import os
import shutil
from typing import Iterable

# Metadati di modulo
__program_name__ = "SIM"
__version__ = "1.5"
__author__ = "LUCA BIMBI"
__date__ = "2025-08-30"

# Fattore di conversione per passare da logaritmo a cents (scala di Ellis)
# 1 ottava = 1200 cents, e log base 10 => si usa 1200 / log(2)
ELLIS_CONVERSION_FACTOR = 1200 / math.log(2)

# Fondamentale di default per la serie subarmonica (impostata da CLI in main)
DEFAULT_SUBHARM_FUND_HZ: float | None = None

# pattern REGEX per determinare funzione e numero tabella in un file esistente
PATTERN = re.compile(r"\bf\s*(\d+)\b")
# \b bonduary; lettera f \s* spazi; (\d+) cifra o cifre

# Tolleranza per confronto floating dei rapporti
RATIO_EPS = 1e-9

def is_fraction_type(value) -> bool:
    """Ritorna True se value è una frazione (fractions.Fraction).

    Utile per distinguere quando l'utente fornisce un rapporto razionale
    invece di un intero per i cents.
    """
    return isinstance(value, Fraction)


def fraction_to_cents(ratio: Fraction) -> int:
    """Converte un rapporto razionale (Fraction) in cents.

    Formula: cents = round(ln(numeratore/denominatore) * 1200 / ln(2)).
    """
    decimal: float = math.log(ratio.numerator / ratio.denominator)
    cents = round(decimal * ELLIS_CONVERSION_FACTOR)
    return cents


def cents_to_fraction(cents: float) -> Fraction:
    """Converte una quantità di cents in un rapporto razionale approssimato.

    Usa l'antilog e poi costruisce una Fraction dal float risultante.
    """
    return Fraction(math.exp(cents / ELLIS_CONVERSION_FACTOR))


def normalize_ratios(ratios: Iterable[float | Fraction], reduce_octave: bool = True) -> list[float]:
    """Normalizza una lista di rapporti:
    - Opzionalmente riduce nell'ottava [1, 2) (default True)
    - Elimina duplicati con tolleranza RATIO_EPS
    - Ordina in modo crescente
    Restituisce una lista di float.
    """
    processed: list[float] = []
    seen: list[float] = []
    for r in ratios:
        v = r
        if reduce_octave:
            v = reduce_to_octave(v)
        v_float = float(v)
        # dedup con tolleranza
        duplicate = False
        for s in seen:
            if abs(v_float - s) <= RATIO_EPS:
                duplicate = True
                break
        if not duplicate:
            seen.append(v_float)
            processed.append(v_float)
    processed.sort()
    return processed


def repeat_ratios(ratios: list[float], span: int, interval_factor: float) -> list[float]:
    """Ripete la serie di rapporti su un certo ambitus moltiplicando per potenze dell'intervallo.

    - ratios: lista base di rapporti (preferibilmente ordinata crescente)
    - span: numero di ripetizioni dell'intervallo (>=1). span=1 ritorna la lista originale.
    - interval_factor: fattore dell'intervallo di ripetizione (es. 2.0 per ottava, oppure exp(cents/...) per ET).

    Restituisce una nuova lista di float con le ripetizioni concatenate mantenendo l'ordine per-blocco.
    """
    try:
        s = int(span)
    except Exception:
        s = 1
    if s <= 1:
        return list(ratios)
    out: list[float] = []
    # Gestione di valori non positivi nell'intervallo: in tal caso non ripetiamo
    if not isinstance(interval_factor, (int, float)) or interval_factor <= 0:
        return list(ratios)
    for k in range(s):
        factor_k = interval_factor ** k
        for r in ratios:
            out.append(float(r) * factor_k)
    return out


def pow_fraction(fr: Fraction | float, k: int) -> Fraction | float:
    """Eleva una Fraction/float ad un esponente intero (anche negativo)."""
    if isinstance(fr, Fraction):
        return fr ** int(k)
    else:
        return float(fr) ** int(k)


def build_natural_ratios(a_max: int, b_max: int, reduce_octave: bool = True) -> list[float]:
    """Genera rapporti del sistema naturale (4:5:6) usando ((3/2)^a)*((5/4)^b).
    a in [-a_max...a_max], b in [-b_max...b_max].
    """
    three_over_two: Fraction = Fraction(3, 2)
    five_over_four: Fraction = Fraction(5, 4)
    vals: list[Fraction | float] = []
    for a in range(-abs(a_max), abs(a_max) + 1):
        for b in range(-abs(b_max), abs(b_max) + 1):
            r = pow_fraction(three_over_two, a) * pow_fraction(five_over_four, b)
            vals.append(r)
    return normalize_ratios(vals, reduce_octave=reduce_octave)


def build_danielou_ratios(full_grid: bool = False, reduce_octave: bool = True) -> list[float]:
    """Genera rapporti del sistema Danielou.
    Formula: ((6/5)^a) * (3/2)^b * (2^c), con c usato per la riduzione nell'ottava.

    Modalità:
    - full_grid=True: genera esattamente i 53 gradi di Danielou seguendo la struttura delle serie descritta:
      a ∈ [-3..3] e, per ciascun a, un numero di quinte ascendenti/descendenti così definito:
        • a = 0 (serie base): 5 quinte discendenti e 5 ascendenti → b ∈ [-5..5]
        • a = 1 (6/5): 3 quinte discendenti e 4 ascendenti → b ∈ [-3..4]
        • a = -1 (5/3): 5 quinte discendenti e 3 ascendenti → b ∈ [-5..3]
        • a = 2 e a = -2: 4 quinte discendenti e 4 ascendenti → b ∈ [-4..4]
        • a = 3 (216/125): solo 4 quinte discendenti (più il grado base) → b ∈ [-4..0]
        • a = -3 (125/108): solo 4 quinte ascendenti (più il grado base) → b ∈ [0..4]
      Tutti i rapporti sono ridotti nell'ottava [1,2), deduplicati e ordinati, poi si prende 1/1 allo step 0,
      i successivi 51 valori e si aggiunge 2/1 come ultimo grado, per un totale di 53.
    - full_grid=False: genera un sottoinsieme dimostrativo (tonica, asse delle quinte, 6/5^k e (5/3)^k per k=1..3).
    """
    six_over_five: Fraction = Fraction(6, 5)
    three_over_two: Fraction = Fraction(3, 2)
    vals: list[Fraction | float] = []

    if full_grid:
        # Mappa "a" → (desc_quinte, asc_quinte)
        series_spec: dict[int, tuple[int, int]] = {
            0: (5, 5),
            1: (3, 4),
            -1: (5, 3),
            2: (4, 4),
            -2: (4, 4),
            3: (4, 0),
            -3: (0, 4),
        }
        for a, (desc_n, asc_n) in series_spec.items():
            # b negativi (quinte discendenti): -desc_n..-1
            for b in range(-desc_n, 0):
                r = pow_fraction(six_over_five, a) * pow_fraction(three_over_two, b)
                vals.append(r)
            # b=0 (centro serie)
            vals.append(pow_fraction(six_over_five, a))
            # b positivi (quinte ascendenti): 1..asc_n
            for b in range(1, asc_n + 1):
                r = pow_fraction(six_over_five, a) * pow_fraction(three_over_two, b)
                vals.append(r)
    else:
        vals.append(Fraction(1, 1))
        # Asse delle quinte (a=0)
        for b in range(-5, 6):
            vals.append(pow_fraction(three_over_two, b))
        # Tre terze minori armoniche successive
        for k in range(1, 4):
            vals.append(pow_fraction(six_over_five, k))
        # Tre seste maggiori armoniche successive
        five_over_three: Fraction = Fraction(5, 3)
        for k in range(1, 4):
            vals.append(pow_fraction(five_over_three, k))

    normalized = normalize_ratios(vals, reduce_octave=reduce_octave)

    if full_grid and reduce_octave:
        # Costruisci lista finale di 53 gradi: 1/1 + 51 valori successivi + 2/1
        base = list(normalized)
        # garantisci 1/1 in testa
        if not base or abs(base[0] - 1.0) > RATIO_EPS:
            base.append(1.0)
            base = sorted(base)
        # prendi i primi 52 in [1,2)
        base_52 = base[:52]
        # dedup fine
        base_52_unique: list[float] = []
        for v in base_52:
            if not base_52_unique or abs(v - base_52_unique[-1]) > RATIO_EPS:
                base_52_unique.append(v)
        # se meno di 52, estendi attingendo dagli elementi rimanenti (già ordinati)
        if len(base_52_unique) < 52 and len(base) > len(base_52_unique):
            for v in base[len(base_52_unique):]:
                if all(abs(v - u) > RATIO_EPS for u in base_52_unique):
                    base_52_unique.append(v)
                    if len(base_52_unique) >= 52:
                        break
        normalized = base_52_unique[:52] + [2.0]
    elif full_grid and reduce_octave and len(normalized) > 53:
        normalized = normalized[:53]
    return normalized


def danielou_from_exponents(a: int, b: int, c: int, reduce_octave: bool = True) -> list[float]:
    """Calcola un singolo rapporto del sistema Danielou dai tre esponenti (a,b,c).

    Formula: r = (6/5)^a * (3/2)^b * (2)^c
    - Se reduce_octave=True, riduce r nell'ottava [1,2).
    Restituisce una lista con un solo valore (float) per compatibilità con le pipeline di export.
    """
    six_over_five: Fraction = Fraction(6, 5)
    three_over_two: Fraction = Fraction(3, 2)
    two: Fraction = Fraction(2, 1)
    r: Fraction | float = (
        pow_fraction(six_over_five, int(a)) *
        pow_fraction(three_over_two, int(b)) *
        pow_fraction(two, int(c))
    )
    r = reduce_to_octave(r) if reduce_octave else r
    return [float(r) if isinstance(r, Fraction) else float(r)]


def reduce_to_octave(value: Fraction | float):
    """Riduce un rapporto nell'ambito di un'ottava [1, 2).

    - Se value è Fraction, mantiene aritmetica razionale.
    - Se value è float, opera in virgola mobile.
    """
    if isinstance(value, Fraction):
        two = Fraction(2, 1)
        while value >= two:
            value /= two
        while value < 1:
            value *= two
        return value
    else:
        # virgola mobile (float)
        while value >= 2.0:
            value /= 2.0
        while value < 1.0:
            value *= 2.0
        return value


def int_or_fraction(value):
    """Argparse type: interpreta una stringa come int oppure come Fraction.

    Esempi validi: "200", "3/2", "7/12". Solleva argparse.ArgumentTypeError
    in caso di valore non interpretabile.
    """
    try:
        return int(value)
    except ValueError:
        try:
            return Fraction(value)
        except ValueError:
            raise argparse.ArgumentTypeError(f"'{value}' non è un intero o una frazione valida.")


def parse_danielou_tuple(value: str) -> tuple[int, int, int]:
    """Argparse type: converte stringhe tipo "a,b,c" in una tupla (a,b,c).

    Formati accettati:
    - "a,b,c" (con spazi opzionali)
    - "a:b:c", "a;b;c"
    - "[a,b,c]", "(a,b,c)", "{a,b,c}"
    - singolo intero "a" → (a, 0, 1)
    """
    if value is None:
        # dovrebbe essere gestito da 'const' su argparse
        return (0, 0, 1)

    s = str(value).strip()
    # Rimuove eventuali parentesi esterne
    if (s.startswith('[') and s.endswith(']')) or \
       (s.startswith('(') and s.endswith(')')) or \
       (s.startswith('{') and s.endswith('}')):
        s = s[1:-1].strip()

    # Normalizza separatori secondari in virgola
    for sep in (':', ';'):
        s = s.replace(sep, ',')

    # Rimuove virgole di coda/duplicati e splitta
    s = s.strip(',')
    parts = [p.strip() for p in s.split(',') if p.strip() != '']

    # Fallback: un singolo intero => (a,0,1)
    if len(parts) == 1:
        try:
            a = int(parts[0])
            return (a, 0, 1)
        except ValueError:
            raise argparse.ArgumentTypeError(
                "Formato non valido per --danielou. Usa 'a,b,c' (es. 0,0,1)."
            )

    if len(parts) != 3:
        raise argparse.ArgumentTypeError(
            "Formato non valido per --danielou. Usa 'a,b,c' (es. 0,0,1)."
        )
    try:
        a, b, c = (int(parts[0]), int(parts[1]), int(parts[2]))
    except ValueError:
        raise argparse.ArgumentTypeError(
            "Esponenti non validi per --danielou (devono essere interi)."
        )
    return (a, b, c)


def ratio_et(index: int, cents: int | Fraction) -> float:
    """Calcola il rapporto della radice per un dato temperamento equabile.

    - index: indice della radice (es. 12 per 12-TET)
    - cents: ampiezza dell'intervallo in cents (di solito 1200) oppure già convertita

    Restituisce il rapporto r tale che r**index = exp(cents / (1200/log(2))).
    """
    decimal_number = math.exp((cents / ELLIS_CONVERSION_FACTOR))  # antilogaritmo
    ratio = decimal_number ** (1 / index)
    return ratio


def note_name_or_frequency(value):
    """Argparse type: prova a interpretare value come float (Hz), altrimenti stringa nota.

    Ritorna float se convertibile, altrimenti la stringa originale (es. "A4").
    """
    try:
        return float(value)
    except (ValueError, TypeError):
        return str(value)

def convert_note_name_to_midi(note_name: str):
    """Segnaposto per la conversione da nome di nota a Hz.

    In futuro: convert_note_name_to_hz(note_name: str, diapason_hz: float = 440.0) -> float
    """
    note_map = {'C': 0, 'D': 2, 'E': 4, 'F': 5, 'G': 7, 'A': 9, 'B': 11}
    note_name = note_name.strip().upper()
    note = note_name[0]
    if note not in note_map:
        raise ValueError(f"Nome nota non valido: {note_name}")
    alteration = 0
    index_octave_number = 1 # indice per il numero di ottava

    if len(note_name) > 1:
        if note_name[1] == '#':
            alteration = 1
            index_octave_number = 2
        elif note_name[1] == 'B':
            alteration = -1
            index_octave_number = 2

    try:
        octave = int(note_name[index_octave_number:]) # converte in int la sottostringa a partire da index_octave_number
    except (ValueError, IndexError):
        raise ValueError(f"Formato ottava non valido in: {note_name}")

    base_note_value = note_map[note]

    midi_value = (octave + 1) * 12 + base_note_value + alteration # +1 allinea MIDI e notazione scientifica (C4=MIDI 60)

    if not (0 <= midi_value <= 127):
        raise ValueError(f"Nota MIDI fuori range (0-127): {midi_value}")

    return midi_value

def convert_midi_to_hz(midi_value: int, diapason_hz: float = 440.0) -> float:
    # 1. calcola quanti semitoni midi_value è distante da A4 (valore 69)
    # 2. converte questo intervallo in semitoni in numero di ottave
    # 3. calcola il fattore per cui la frequenza deve essere moltiplicata
    # 4. moltiplica per il diapason specificato
    hz = diapason_hz * (2 ** ((float(midi_value) - 69) / 12))
    return hz

def file_exists(file_path: str) -> bool:
    """Verifica se il file esiste."""
    if os.path.exists(file_path):
        print(f"File esistente: {file_path}")
        return True
    else:
        print(f"File non esistente: {file_path}")
        return False

def parse_file(file_name: str) -> int:
    """Analizza il file alla ricerca di pattern 'f N' e ritorna il massimo N trovato.

    Se il file non esiste o non ci sono match, ritorna 0.
    """
    max_num: int | None = None
    try:
        with open(file_name, "r") as file:
            for line in file:
                for match in PATTERN.finditer(line):
                    num = int(match.group(1))
                    max_num = num if max_num is None else max(max_num, num)
    except FileNotFoundError:
        print(f"Errore: file non trovato: {file_name}")
        return 0
    return 0 if max_num is None else max_num


def write_frequency_table(output_base: str, rows: list[tuple[int, float]]) -> None:
    """[Deprecato] Scrive frequenze in '<output_base>.txt'. Mantenuto per compatibilità temporanea."""
    path = f"{output_base}.txt"
    try:
        mode = "a+" if file_exists(path) else "w"
        with open(path, mode) as file:
            for step_idx, freq in rows:
                file.write(f"{freq:.2f} Hz  Step {step_idx}\n")
        print(f"Tabella salvata in {path}")
    except Exception as e:
        print(f"Errore scrittura file: {e}")


def ensure_midi_fit(ratios: list[float], basekey: int, prefer_truncate: bool) -> tuple[list[float], int]:
    """Ensures the mapping of 'ratios' starting at 'basekey' fits the MIDI range [0..127].

    - If prefer_truncate=False (default behavior): try to adapt the effective basekey so the whole series fits.
      If impossible (len(ratios) > 128), truncate to 128 and print a WARNING.
    - If prefer_truncate=True: clamp basekey to [0,127] and truncate the series to fit.

    Returns: (ratios_eff, basekey_eff)
    """
    n = len(ratios)
    # Clamp provided basekey to a reasonable integer
    try:
        bk = int(basekey)
    except Exception:
        bk = 0
    # If too many notes, we can only keep 128 max
    if n > 128:
        if not prefer_truncate:
            print(f"WARNING: numero di passi ({n}) eccede 128. Adattamento impossibile: verranno mantenuti solo i primi 128 e basekey verrà adattata.")
            prefer_truncate = True
        n = 128
        ratios = list(ratios[:n])
    if prefer_truncate:
        # Clamp basekey, then truncate window to fit
        bk = max(0, min(127, bk))
        max_len = 128 - bk
        if len(ratios) > max_len:
            print(f"WARNING: serie eccede il limite MIDI con basekey={bk}. Verranno troncati {len(ratios)-max_len} passi oltre la nota 127.")
            ratios = list(ratios[:max_len])
        return ratios, bk
    else:
        # Adapt basekey so that bk in [0, 127 - (n-1)]
        allowed_min = 0
        allowed_max = 127 - (n - 1)
        if allowed_max < allowed_min:
            # Should not happen because n<=128 here, but guard anyway
            allowed_max = allowed_min
        bk_eff = bk
        if bk < allowed_min or bk > allowed_max:
            # prefer to keep close to user value but clamp into range
            bk_eff = max(allowed_min, min(allowed_max, bk))
            print(f"WARNING: basekey adattata da {bk} a {bk_eff} per includere tutti i {n} passi entro MIDI 0..127.")
        return ratios, bk_eff

def write_cpstun_table(output_base: str, ratios: list[float], basekey: int, basefrequency: float, interval_value: float | None = None) -> tuple[int, bool]:
    """Crea o appende una tabella cpstun in un file Csound .csd.

    Formato GEN -2 atteso da cpstun:
    fN 0 size -2  numgrades interval basefreq basekey  r1 r2 r3 ...

    - numgrades: numero di gradi nell'intervallo (len(ratios))
    - interval: intervallo di ripetizione del sistema.
      • Se interval_value > 0 è fornito, viene usato direttamente (es. 2 per ottava, 1.5 per quinta, ecc.).
      • Altrimenti, inferenza di fallback: 2.0 se tutti i rapporti sono nell'ottava [1,2] (inclusiva)
        e c'è almeno un valore strettamente < 2; altrimenti 0.0.
    - basefreq: frequenza assoluta associata a basekey (basenote in Hz)
    - basekey: nota MIDI a cui si riferisce basefreq
    - r1..rn: rapporti unitari

    Ritorna una tupla (fnum, existed_before) dove:
    - fnum è il numero della tabella appena scritta;
    - existed_before indica se il file .csd esisteva già prima della scrittura.
    """
    csd_path = f"{output_base}.csd"
    skeleton = (
        "<CsoundSynthesizer>\n"
        "<CsOptions>\n\n</CsOptions>\n"
        "<CsInstruments>\n\n</CsInstruments>\n"
        "<CsScore>\n\n</CsScore>\n"
        "</CsoundSynthesizer>\n"
    )

    existed_before = file_exists(csd_path)
    if not existed_before:
        # crea scheletro .csd
        try:
            with open(csd_path, "w") as f:
                f.write(skeleton)
        except Exception as e:
            print(f"Errore creazione file CSD: {e}")
            return 0, existed_before

    # leggi contenuto corrente
    try:
        with open(csd_path, "r", encoding="utf-8") as f:
            content = f.read()
    except Exception as e:
        print(f"Errore lettura CSD: {e}")
        return 0, existed_before

    # calcola prossimo numero di tabella
    f_max = parse_file(csd_path)
    fnum = (f_max or 0) + 1

    # prima di scrivere, ordina i rapporti in ordine crescente (requisito utente)
    try:
        ratios_sorted = sorted(float(r) for r in ratios)
    except Exception:
        ratios_sorted = [float(r) for r in ratios]

    # determina header cpstun
    numgrades = int(len(ratios_sorted))
    # intervallo di ripetizione: usa interval_value se fornito; consenti esplicitamente 0 per non ripetibilità.
    if interval_value is not None and isinstance(interval_value, (int, float)):
        if float(interval_value) <= 0:
            interval = 0.0
        else:
            interval = float(interval_value)
    else:
        try:
            rmin = min(ratios_sorted) if ratios_sorted else 1.0
            rmax = max(ratios_sorted) if ratios_sorted else 1.0
        except Exception:
            rmin, rmax = 1.0, 1.0
        # accetta [1,2] inclusivo come finestra d'ottava
        eps = 1e-9
        interval = 2.0 if (rmin >= 1.0 - eps and rmax <= 2.0 + eps) else 0.0

    # compone la lista dati: header + ratios
    data_list: list[str] = []
    data_list.append(str(numgrades))                # numgrades (int)
    data_list.append(f"{float(interval):.10g}")    # interval
    data_list.append(f"{float(basefrequency):.10g}")  # basefreq
    data_list.append(str(int(basekey)))             # basekey
    data_list.extend(f"{r:.10f}" for r in ratios_sorted)

    # dimensione tabella = numero di dati forniti a GEN -2
    size = len(data_list)

    # Costruisce dinamicamente righe di commento allineate in modo che le etichette cadano sotto l'inizio dei token
    prefix = f"f {fnum} 0 {size} -2 "
    token_strs = list(data_list)
    # Calcola la colonna iniziale (0-based) di ciascun token nella riga f
    positions: list[int] = []
    col = len(prefix)
    for t in token_strs:
        positions.append(col)
        col += len(t) + 1  # token più spazio successivo

    def build_aligned_comment_line(label_map: list[tuple[int, str]]) -> str:
        # label_map: lista di (token_index, label_text) da posizionare su questa linea
        line = ";"  # il commento inizia con il punto e virgola
        base_offset = 1  # tiene conto del ';' che occupa la colonna 0
        for idx, text in label_map:
            target = base_offset + (positions[idx] if 0 <= idx < len(positions) else len(prefix))
            if target < len(line):
                # garantisce almeno uno spazio tra token ed etichetta
                target = len(line) + 1
            line += " " * (target - len(line)) + text
        return line + "\n"

    # Definisce quale etichetta finisce su quale linea
    line1 = build_aligned_comment_line([
        (0, "numgrades"),
        (2, "basefreq"),
        (4, "rapporti-di-intonazione .......")
    ])
    line2 = build_aligned_comment_line([
        (1, "interval"),
        (3, "basekey")
    ])
    header_comment = (
        line1 +
        line2 +
        f"; tabella cpstun generata automaticamente | basekey={basekey} basefrequency={basefrequency:.6f}Hz\n"
    )
    f_line = prefix + " ".join(data_list) + "\n"

    # inserisci prima di </CsScore>
    insert_marker = "</CsScore>"
    idx = content.rfind(insert_marker)
    if idx == -1:
        # struttura mancante: appende una sezione <CsScore>
        content = (
            content +
            "\n<CsScore>\n" + header_comment + f_line + "</CsScore>\n"
        )
    else:
        content = content[:idx] + header_comment + f_line + content[idx:]

    try:
        with open(csd_path, "w", encoding="utf-8") as f:
            f.write(content)
        print(f"Tabella cpstun (f {fnum}) salvata in {csd_path}")
    except Exception as e:
        print(f"Errore scrittura CSD: {e}")
        return (fnum, existed_before)

    return (fnum, existed_before)

def write_tun_file(output_base: str, ratios: list[float], basekey: int, basefrequency: float) -> None:
    """Esporta un file .tun (AnaMark TUN) con sezione [Exact Tuning] (Note 0..127 in Hz).

    - Mappa i passi generati su note MIDI consecutive a partire da basekey.
    - Le note non coperte dai passi vengono riempite con 12TET relativo a basekey.
    - basefrequency è la frequenza assoluta della basekey (basenote convertita in Hz).
    """
    tun_path = f"{output_base}.tun"
    lines: list[str] = []
    lines.append("[Tuning]")
    lines.append("FormatVersion=1")
    lines.append(f"Name=Generated by {__program_name__} {__version__}")
    lines.append("")
    lines.append("[Exact Tuning]")
    lines.append(f"; basekey={basekey} basefrequency={basefrequency:.6f}Hz")

    # Precomputa 12TET fallback relativo a basefrequency
    def tet_freq(offset_semitones: int) -> float:
        return basefrequency * (2.0 ** (offset_semitones / 12.0))

    for n in range(128):
        if 0 <= basekey <= 127 and (basekey <= n < basekey + len(ratios)):
            idx = n - basekey
            if 0 <= idx < len(ratios):
                freq = basefrequency * float(ratios[idx])
            else:
                freq = tet_freq(n - basekey)
        else:
            freq = tet_freq(n - basekey)
        lines.append(f"Note {n}={freq:.10f} Hz")

    try:
        with open(tun_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")
        print(f"File .tun salvato in {tun_path}")
    except Exception as e:
        print(f"Errore scrittura .tun: {e}")


def export_system_tables(output_base: str, ratios: list[float], basekey: int, basenote_hz: float) -> None:
    """Esporta tabelle del sistema generato (Hz / ratio / MIDI) in testo ed Excel.

    Crea:
    - {output_base}_system.txt (colonne allineate a larghezza fissa)
    - {output_base}_system.xlsx (se openpyxl disponibile)

    Requisito: gli intervalli del sistema custom sono ordinati per Hz crescente.
    """
    # Prepara righe calcolando Hz e mantenendo riferimento a step originale
    computed = []  # (hz, step, midi, ratio)
    for i, r in enumerate(ratios):
        midi = basekey + i
        hz = basenote_hz * float(r)
        computed.append((hz, i, midi, float(r)))
    # Ordina per Hz crescente
    computed.sort(key=lambda t: t[0])

    # Testo - costruzione tabella a colonne allineate
    txt_path = f"{output_base}_system.txt"
    try:
        headers = ["Step", "MIDI", "Ratio", "Hz"]
        rows = []
        for row_idx, (hz, i, midi, r) in enumerate(computed):
            rows.append([
                str(row_idx),
                str(basekey + row_idx),
                f"{r:.10f}",
                f"{hz:.6f}",
            ])
        # Calcola larghezze massime per colonna
        widths = [len(h) for h in headers]
        for row in rows:
            for c, val in enumerate(row):
                if len(val) > widths[c]:
                    widths[c] = len(val)
        # Funzione per formattare una riga con padding a destra
        def fmt(vals: list[str]) -> str:
            return "  ".join(val.ljust(widths[i]) for i, val in enumerate(vals))
        with open(txt_path, "w", encoding="utf-8") as f:
            f.write(fmt(headers) + "\n")
            for row in rows:
                f.write(fmt(row) + "\n")
        print(f"Esportato: {txt_path}")
    except Exception as e:
        print(f"Errore scrittura {txt_path}: {e}")

    # Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "System"
        headers = ["Step", "MIDI", "Ratio", "Hz"]
        ws.append(headers)
        header_fill = PatternFill(start_color="FFDDDDDD", end_color="FFDDDDDD", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for row_idx, (hz, i, midi, r) in enumerate(computed):
            ws.append([row_idx, basekey + row_idx, r, hz])
        xlsx_path = f"{output_base}_system.xlsx"
        wb.save(xlsx_path)
        print(f"Esportato: {xlsx_path}")
    except ImportError:
        print("openpyxl non installato: export Excel saltato per _system.xlsx")
    except Exception as e:
        print(f"Errore export Excel _system.xlsx: {e}")


def export_comparison_tables(output_base: str, ratios: list[float], basekey: int, basenote_hz: float,
                             diapason_hz: float, compare_fund_hz: float | None = None, tet_align: str = "same",
                             subharm_fund_hz: float | None = None) -> None:
    """Esporta tabelle di confronto con 12TET, serie armonica e serie subarmonica.

    Crea:
    - {output_base}_compare.txt
    - {output_base}_compare.xlsx (se openpyxl disponibile)
    Colonne TXT: Step, MIDI, Ratio, Custom_Hz, Harmonic_Hz, DeltaHz_Harm, Subharm_Hz, DeltaHz_Sub, TET_Hz, TET_Note, DeltaHz_TET
    In Excel abbiniamo valori vicini affiancandoli (Custom accanto a Harmonic/Subharmonic) e li coloriamo
    con colori distinti: Custom in rosso, Harmonic in verde, Subharm in giallo, TET in blu. Se |Delta| < 17 Hz
    tra Custom e Harmonic, evidenziamo entrambi in grassetto.

    Parametri aggiuntivi:
    - compare_fund_hz: fondamentale per la serie armonica e per l'ancoraggio del 12TET.
      Se None, si usa basenote_hz.
    - tet_align: "same" (il 12TET parte esattamente da compare_fund_hz) oppure "nearest"
      (per ogni Custom si prende la nota 12TET più vicina nel reticolo ancorato a compare_fund_hz).
    - subharm_fund_hz: fondamentale per la serie subarmonica. Se None, si usa A4 del diapason corrente (es. 440 Hz).

    Requisiti aggiunti:
    - Ordina gli intervalli del sistema custom per Hz crescente nelle comparazioni.
    - Le armoniche non devono superare 10 kHz e le subarmoniche non devono scendere sotto 16 Hz.
    """
    proximity_thr = 17.0  # Hz per considerare "vicini"
    MAX_HARM_HZ = 10000.0
    MIN_SUB_HZ = 16.0

    # Conversione freq -> nome nota 12TET rispetto al diapason (A4 = diapason_hz)
    def freq_to_note_name(freq: float, a4_hz: float) -> str:
        try:
            if not (freq > 0 and a4_hz > 0):
                return ""
            midi = int(round(69 + 12.0 * math.log2(freq / a4_hz)))
        except Exception:
            return ""
        midi = max(0, min(127, midi))
        names = ["C", "C#", "D", "D#", "E", "F", "F#", "G", "G#", "A", "A#", "B"]
        name = names[midi % 12]
        octave = (midi // 12) - 1
        return f"{name}{octave}"

    base_cmp = compare_fund_hz if compare_fund_hz is not None else basenote_hz
    sub_base = (
        subharm_fund_hz
        if subharm_fund_hz is not None
        else (DEFAULT_SUBHARM_FUND_HZ if DEFAULT_SUBHARM_FUND_HZ is not None else diapason_hz)
    )  # A4 del diapason o valore globale impostato

    def tet_from_base(step_i: int, custom_val: float | None = None) -> float:
        if tet_align == "nearest" and custom_val is not None and custom_val > 0 and base_cmp > 0:
            # trova il grado 12TET più vicino a custom_val nel reticolo ancorato a base_cmp
            n = round(12.0 * math.log2(custom_val / base_cmp))
            return base_cmp * (2.0 ** (n / 12.0))
        # default: stesso indice i dalla fondamentale
        return base_cmp * (2.0 ** (step_i / 12.0))

    # Prepara elenco calcolato e ordina per Hz crescente
    computed = []  # (custom_hz, i, midi, ratio)
    for i, r in enumerate(ratios):
        midi = basekey + i
        custom_hz = basenote_hz * float(r)
        computed.append((custom_hz, i, midi, float(r)))
    computed.sort(key=lambda t: t[0])

    # Costruisci sequenze crescenti per colonne di confronto
    # Custom già in ordine crescente in 'computed'
    custom_list = computed  # alias
    n_rows = len(custom_list)

    # Serie armonica crescente (1,2,3,...) fino a 10 kHz
    harm_vals: list[float] = []
    n = 1
    while True:
        val = base_cmp * n
        if val > MAX_HARM_HZ:
            break
        harm_vals.append(val)
        n += 1

    # Serie subarmonica: dal valore più basso >= MIN_SUB_HZ fino al generatore (sub_base)
    sub_desc: list[float] = []  # decrescente mentre n cresce
    m = 1
    while True:
        val = sub_base / m
        if val < MIN_SUB_HZ:
            break
        sub_desc.append(val)
        m += 1
    sub_vals = list(reversed(sub_desc))  # crescente dai più bassi fino al generatore

    # 12TET crescente: trova il primo grado >= min custom e poi sale a semitoni
    min_custom = custom_list[0][0] if n_rows > 0 else base_cmp
    # calcola offset in semitoni relativo a base_cmp
    if min_custom > 0 and base_cmp > 0:
        offset = 12.0 * math.log2(min_custom / base_cmp)
    else:
        offset = 0.0
    if tet_align == "nearest":
        start_n = int(round(offset))
    else:
        start_n = int(math.ceil(offset))
    tet_vals: list[float] = []
    for k in range(max(n_rows, 0)):
        tet_vals.append(base_cmp * (2.0 ** ((start_n + k) / 12.0)))

    # Applica cutoff alle subarmoniche: escludi tutte quelle sotto la prima Custom (oltre a 16 Hz)
    # Soglia = max(MIN_SUB_HZ, min_custom)
    cutoff_sub = max(MIN_SUB_HZ, min_custom)
    sub_vals = [v for v in sub_vals if v >= cutoff_sub]

    # Funzione di allineamento: distribuisce una sequenza crescente nelle finestre [c_i, c_{i+1})
    def align_sequence(seq: list[float], customs: list[float]) -> list[float | None]:
        out: list[float | None] = [None] * len(customs)
        p = 0
        for i in range(len(customs)):
            low = customs[i]
            high = customs[i + 1] if i + 1 < len(customs) else float('inf')
            if p < len(seq) and low <= seq[p] < high:
                out[i] = seq[p]
                p += 1
            else:
                out[i] = None
        return out

    # Costruiamo liste Custom in Hz (crescente)
    custom_hz_list = [c[0] for c in custom_list]

    # Prepara sequenze allineate
    harm_aligned = align_sequence(harm_vals, custom_hz_list)
    sub_aligned = align_sequence(sub_vals, custom_hz_list)
    tet_aligned = align_sequence(tet_vals, custom_hz_list)

    # Testo (mettiamo Custom e Harmonic/Subharmonic affiancati) con colonne allineate a larghezza fissa
    txt_path = f"{output_base}_compare.txt"
    try:
        headers = [
            "Step", "MIDI", "Ratio",
            "Custom_Hz", "Harmonic_Hz", "DeltaHz_Harm",
            "Subharm_Hz", "DeltaHz_Sub",
            "TET_Hz", "TET_Note", "DeltaHz_TET",
        ]
        rows: list[list[str]] = []
        for row_i, (custom_hz, i, midi, r) in enumerate(custom_list):
            harm_val = harm_aligned[row_i]
            sub_val = sub_aligned[row_i]
            # 12TET per-row assignment: always nearest note anchored to base_cmp
            n = round(12.0 * math.log2(custom_hz / base_cmp)) if (custom_hz > 0 and base_cmp > 0) else 0
            tet_val_row = base_cmp * (2.0 ** (n / 12.0))

            # Harmonic
            harm_str = ""
            d_har_str = ""
            approx = ""
            if harm_val is not None:
                d_har = custom_hz - harm_val
                harm_str = f"{harm_val:.6f}"
                d_har_str = f"{d_har:.6f}"
                approx = "≈" if abs(d_har) < proximity_thr else ""

            # Subharmonic
            sub_str = ""
            d_sub_str = ""
            if sub_val is not None:
                d_sub = custom_hz - sub_val
                sub_str = f"{sub_val:.6f}"
                d_sub_str = f"{d_sub:.6f}"

            # 12TET
            tet_note = ""
            d_tet_str = ""
            tet_str = ""
            if tet_val_row is not None and isinstance(tet_val_row, (int, float)) and math.isfinite(tet_val_row) and tet_val_row > 0:
                tet_str = f"{tet_val_row:.6f}"
                tet_note = freq_to_note_name(tet_val_row, diapason_hz)
                d_tet = custom_hz - tet_val_row
                d_tet_str = f"{d_tet:.6f}"

            rows.append([
                str(row_i),
                str(basekey + row_i),
                f"{r:.10f}",
                f"{custom_hz:.6f}{approx}",
                f"{harm_str}{approx if harm_str else ''}",
                f"{d_har_str}",
                f"{sub_str}",
                f"{d_sub_str}",
                f"{tet_str}",
                f"{tet_note}",
                f"{d_tet_str}",
            ])
        # Calcola larghezze massime per colonna
        widths = [len(h) for h in headers]
        for row in rows:
            for c, val in enumerate(row):
                if len(val) > widths[c]:
                    widths[c] = len(val)
        def fmt(vals: list[str]) -> str:
            return "  ".join(val.ljust(widths[i]) for i, val in enumerate(vals))
        with open(txt_path, "w", encoding="utf-8") as f:
            f.write(fmt(headers) + "\n")
            for row in rows:
                f.write(fmt(row) + "\n")
        print(f"Esportato: {txt_path}")
    except Exception as e:
        print(f"Errore scrittura {txt_path}: {e}")

    # Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "Compare"
        headers = [
            "Step", "MIDI", "Ratio",
            "Custom_Hz", "Harmonic_Hz", "|DeltaHz_Harm|",
            "Subharm_Hz", "|DeltaHz_Sub|",
            "TET_Hz", "TET_Note", "|DeltaHz_TET|",
        ]
        ws.append(headers)
        header_fill = PatternFill(start_color="FFCCE5FF", end_color="FFCCE5FF", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        # Riempimenti per le serie
        fill_custom = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid")  # rosso chiaro
        fill_harm = PatternFill(start_color="FFCCFFCC", end_color="FFCCFFCC", fill_type="solid")    # verde chiaro
        fill_sub = PatternFill(start_color="FFFFFFCC", end_color="FFFFFFCC", fill_type="solid")      # giallo chiaro
        fill_tet = PatternFill(start_color="FFCCE5FF", end_color="FFCCE5FF", fill_type="solid")      # blu chiaro

        for row_i, (custom_hz, i, midi, r) in enumerate(custom_list):
            harm_val = harm_aligned[row_i]
            sub_val = sub_aligned[row_i]
            # 12TET per-row assignment: always nearest note anchored to base_cmp (same logic as TXT)
            n = round(12.0 * math.log2(custom_hz / base_cmp)) if (custom_hz > 0 and base_cmp > 0) else 0
            tet_val_row = base_cmp * (2.0 ** (n / 12.0))

            harm_cell = None
            d_har_cell = None
            is_close = False
            if harm_val is not None:
                d_har = custom_hz - harm_val
                harm_cell = harm_val
                d_har_cell = abs(d_har)
                is_close = abs(d_har) < proximity_thr

            sub_cell = None
            d_sub_cell = None
            if sub_val is not None:
                d_sub = custom_hz - sub_val
                sub_cell = sub_val
                d_sub_cell = abs(d_sub)

            tet_cell = None
            tet_note = ""
            d_tet_cell = None
            if tet_val_row is not None and isinstance(tet_val_row, (int, float)) and math.isfinite(tet_val_row) and tet_val_row > 0:
                tet_cell = tet_val_row
                tet_note = freq_to_note_name(tet_val_row, diapason_hz)
                d_tet_cell = abs(custom_hz - tet_val_row)

            ws.append([row_i, basekey + row_i, r, custom_hz, harm_cell, d_har_cell, sub_cell, d_sub_cell, tet_cell, tet_note, d_tet_cell])
            row = ws.max_row
            # Applica colorazione per serie
            ws.cell(row=row, column=4).fill = fill_custom
            ws.cell(row=row, column=5).fill = fill_harm
            ws.cell(row=row, column=7).fill = fill_sub
            ws.cell(row=row, column=9).fill = fill_tet
            ws.cell(row=row, column=10).fill = fill_tet
            # Evidenziazione disabilitata nel formato Excel su richiesta utente:
            # niente grassetto condizionale per prossimità (< 17 Hz).
        xlsx_path = f"{output_base}_compare.xlsx"
        wb.save(xlsx_path)
        print(f"Esportato: {xlsx_path}")
    except ImportError:
        print("openpyxl non installato: export Excel saltato per _compare.xlsx")
    except Exception as e:
        print(f"Errore export Excel _compare.xlsx: {e}")


def print_step_hz_table(ratios: list[float], basenote_hz: float) -> None:
    """Stampa su console una tabella multi-colonna con coppie Step/Hz che riempiono lo schermo.

    - Step numerato da 1.
    - Hz con due decimali.
    - Disposizione a più colonne in base alla larghezza del terminale.
    """
    # Prepara dati Step/Hz
    pairs: list[tuple[str, str]] = []
    for i, r in enumerate(ratios):
        hz = basenote_hz * float(r)
        pairs.append((str(i + 1), f"{hz:.2f}"))

    # Calcola larghezze per formattazione
    step_w = max(len("Step"), max((len(p[0]) for p in pairs), default=1))
    hz_w = max(len("Hz"), max((len(p[1]) for p in pairs), default=1))
    cell_w = step_w + 1 + hz_w  # "Step" + spazio + "Hz"
    gap = 3  # spazi tra le celle

    # Determina larghezza terminale
    try:
        term_w = shutil.get_terminal_size(fallback=(80, 24)).columns
    except Exception:
        term_w = 80

    # Numero colonne che entrano nella larghezza disponibile
    if cell_w >= term_w:
        cols = 1
    else:
        # massimo cols tale che cols*cell_w + (cols-1)*gap <= term_w
        cols = max(1, (term_w + gap) // (cell_w + gap))

    n = len(pairs)
    header_cell = "Step".ljust(step_w) + " " + "Hz".ljust(hz_w)
    if n == 0:
        print(header_cell)
        return

    rows_count = math.ceil(n / cols)

    # Stampa intestazione ripetuta per ogni colonna
    print((" " * gap).join([header_cell] * cols))

    # Stampa righe in ordine colonna-major per bilanciare l'altezza
    for r in range(rows_count):
        line_cells: list[str] = []
        for c in range(cols):
            idx = c * rows_count + r
            if idx < n:
                s, h = pairs[idx]
                cell = s.rjust(step_w) + " " + h.rjust(hz_w)
                line_cells.append(cell)
        if line_cells:
            print((" " * gap).join(line_cells))


def main():
    """Punto di ingresso CLI.

    - Mostra l'help di argparse quando non vengono passati argomenti.
    - Imposta e valida gli argomenti della riga di comando.
    - Esegue il calcolo del rapporto della radice per il temperamento richiesto.
    """
    parser = argparse.ArgumentParser(
        description=(
            "SIM. Assistente software per i sistemi di intonazione musicale. "
            "Generatore di tabelle/rapporti per cpstun in Csound, file TUN e tabelle di comparazione con 12TET. "
            f"Versione {__version__}"
        ),
        epilog="Copyright (C) 2025 Luca Bimbi. Questo progetto è rilasciato sotto licenza MIT. Vedi il file LICENSE"
    )

    # Versione del programma
    parser.add_argument("-v", "--version", action="version", version=f"%(prog)s {__version__}",
                        help="Visualizza la versione del programma")

    # Diapason di riferimento in Hz (stringa per lasciare libertà input; conversione demandata altrove)
    parser.add_argument("--diapason", type=float, default=440.0,
                        help="Diapason di riferimento in Hz (default: 440)")

    # Nota base per tabella cpstun (MIDI key), default 60 (C4)
    parser.add_argument(
        "--basekey", default=60, type=int,
        help="Nota base (MIDI key) per tabella cpstun (default: 60)"
    )

    # Nota/frequenza di riferimento
    parser.add_argument(
        "--basenote", default="C4", type=note_name_or_frequency,
        help="Nota di riferimento (es. 'A4', 'F#2', 'Ab3') o frequenza in Hz"
    )

    # Temperamento equabile: indice della radice e ampiezza in cents (o frazione convertibile)
    parser.add_argument(
        "--et", nargs=2, default=(12, 200), metavar=("INDEX", "INTERVAL"), type=int_or_fraction,
        help=(
            "Temperamento equabile definibile dall'utente. "
            "Richiede: indice della radice e valore dell'intervallo in cents (o frazione)."
        ),
    )

    # Sistema geometrico basato su un generatore (es. 3/2) e numero di passi
    parser.add_argument(
        "--geometric", nargs=2, metavar=("GEN", "STEPS"),
        help=(
            "Sistema geometrico su generatore (es. 3/2) e passi. "
            "Esempio: --geometric 3/2 12"
        ),
    )
    # Sistema naturale (4:5:6)
    parser.add_argument(
        "--natural", nargs=2, type=int, metavar=("A_MAX", "B_MAX"),
        help=(
            "Sistema naturale 4:5:6 con ((3/2)^a)*((5/4)^b), a∈[-A_MAX..A_MAX], b∈[-B_MAX..B_MAX]"
        ),
    )
    # Sistema Danielou
    parser.add_argument(
        "--danielou", action="append", type=parse_danielou_tuple, default=None,
        help=(
            "Sistema Danielou (manuale): specifica una terna 'a,b,c' (es. 1,2,-1). "
            "Ripeti l'opzione per più terne. Genera solo i rapporti indicati. Usa --danielou-all per la griglia completa."
        )
    )
    parser.add_argument(
        "--danielou-all", action="store_true",
        help="Usa la griglia completa Danielou (a=-3..3, b=-5..5) per ottenere fino a 53 rapporti"
    )
    parser.add_argument(
        "--no-reduce", action="store_true",
        help="Non ridurre i rapporti all'interno dell'ottava (default: riduci)",
    )
    # Ambitus/Span di ripetizione dell'intervallo generato
    parser.add_argument(
        "--span", "--ambitus", dest="span", type=int, default=1,
        help="Numero di ripetizioni dell'intervallo base del sistema (default: 1). Es.: --span 3"
    )
    parser.add_argument(
        "--interval-zero", action="store_true",
        help="Imposta interval=0 nella tabella cpstun (.csd) per non ripetibilità: in questo caso l'ambitus (--span) è considerato e si elencano tanti ratio quanti sono gli step × span",
    )
    parser.add_argument(
        "--export-tun", action="store_true",
        help="Esporta anche un file .tun (AnaMark TUN) con [Exact Tuning] su 128 note",
    )

    # Opzioni per il confronto (serie armonica, subarmonica e 12TET)
    parser.add_argument(
        "--compare-fund", nargs="?", type=note_name_or_frequency, default="basenote", const="basenote",
        help="Fondamentale per il confronto armonico (nota es. 'A4' o frequenza in Hz). Default: basenote (se usato senza argomento prende basenote)"
    )
    parser.add_argument(
        "--compare-tet-align", choices=["same", "nearest"], default="same",
        help="Allineamento 12TET nel confronto: 'same' dalla fondamentale, 'nearest' nota 12TET più vicina alla Custom"
    )
    parser.add_argument(
        "--subharm-fund", type=note_name_or_frequency, default="A5",
        help="Fondamentale per la serie subarmonica (nota es. 'A4' o frequenza in Hz). Default: A5"
    )
    # Controllo mappatura MIDI
    parser.add_argument(
        "--midi-truncate", action="store_true",
        help="Forza il troncamento delle altezze che eccedono il range MIDI 0..127 invece di adattare la basekey"
    )

    # File di output
    parser.add_argument("output_file", help="File di output con la tabella di accordatura")

    # Se non sono stati passati argomenti, mostra l'help e termina
    if len(sys.argv) == 1:
        parser.print_help()
        return

    args = parser.parse_args()

    # Validazione basenote
    if isinstance(args.basenote, (int, float)) and args.basenote < 0:
        print("Nota di riferimento non valida.")
        return

    # Validazione span
    if args.span is None or args.span < 1:
        print("Valore non valido per --span/--ambitus: impostato a 1.")
        args.span = 1

    # Determina la frequenza di base in Hz: se float, è già Hz; altrimenti sarebbe da convertire
    if isinstance(args.basenote, float):
        basenote = args.basenote
    else:
        # conversione del nome nota, utilizzando il diapason
        basenote_midi = convert_note_name_to_midi(args.basenote)
        basenote = convert_midi_to_hz(basenote_midi, args.diapason)

    # Determina la fondamentale per il confronto (serie armonica e ancoraggio 12TET)
    # Usa default= di argparse: "basenote" indica di usare la basenote calcolata
    cf = args.compare_fund
    if isinstance(cf, str) and cf.lower() == "basenote":
        compare_fund_hz = basenote
    elif isinstance(cf, float):
        compare_fund_hz = cf
    else:
        # è un nome di nota: convertilo con il diapason corrente
        cf_midi = convert_note_name_to_midi(cf)
        compare_fund_hz = convert_midi_to_hz(cf_midi, args.diapason)

    # Determina la fondamentale per la serie subarmonica (default via argparse: "A5")
    sf = args.subharm_fund
    if isinstance(sf, float):
        subharm_fund_hz = sf
    elif sf is None:
        # retrocompatibilità: se assente, usa A4 (diapason)
        subharm_fund_hz = float(args.diapason)
    else:
        sh_midi = convert_note_name_to_midi(str(sf))
        subharm_fund_hz = convert_midi_to_hz(sh_midi, args.diapason)

    # Imposta la fondamentale subarmonica di default per le esportazioni
    global DEFAULT_SUBHARM_FUND_HZ
    DEFAULT_SUBHARM_FUND_HZ = subharm_fund_hz

    # Sistema naturale (4:5:6): priorità alta rispetto a geometric/ET
    if args.natural:
        try:
            a_max = int(args.natural[0])
            b_max = int(args.natural[1])
        except (TypeError, ValueError):
            print("Valori non validi per --natural (attesi interi A_MAX B_MAX).")
            return
        if a_max < 0 or b_max < 0:
            print("A_MAX e B_MAX devono essere >= 0 per --natural.")
            return
        ratios = build_natural_ratios(a_max, b_max, reduce_octave=(not args.no_reduce))
        # Applica ambitus/span su ottave per stampa e confronti
        ratios_spanned = repeat_ratios(ratios, args.span, 2.0)
        # MIDI fit (adapt or truncate) per export/stampa
        ratios_eff, basekey_eff = ensure_midi_fit(ratios_spanned, args.basekey, args.midi_truncate)
        print_step_hz_table(ratios_eff, basenote)

        # Preparazione dati per CSD:
        # - di default ignora ambitus (tabella ripetibile tramite 'interval')
        # - con --interval-zero considera l'ambitus: tabella non ripetibile con tutti gli step ripetuti
        if getattr(args, "interval_zero", False):
            csd_input = ratios_spanned
            csd_interval = 0.0
        else:
            csd_input = ratios
            csd_interval = 2.0
        csd_ratios, csd_basekey = ensure_midi_fit(csd_input, args.basekey, args.midi_truncate)
        fnum, existed = write_cpstun_table(args.output_file, csd_ratios, csd_basekey, basenote, csd_interval)

        export_base = args.output_file if not existed else f"{args.output_file}_{fnum}"
        export_system_tables(export_base, ratios_eff, basekey_eff, basenote)
        try:
            diapason_hz = float(args.diapason)
        except TypeError:
            diapason_hz = 440.0
        export_comparison_tables(export_base, ratios_eff, basekey_eff, basenote, diapason_hz,
                                 compare_fund_hz=compare_fund_hz, tet_align=args.compare_tet_align,
                                 subharm_fund_hz=subharm_fund_hz)
        if args.export_tun:
            write_tun_file(export_base, ratios_eff, basekey_eff, basenote)
        return

    # Sistema Danielou: esponenti opzionali via --danielou oppure griglia completa via --danielou-all
    if args.danielou_all:
        ratios = build_danielou_ratios(full_grid=True, reduce_octave=(not args.no_reduce))
        ratios_spanned = repeat_ratios(ratios, args.span, 2.0)
        ratios_eff, basekey_eff = ensure_midi_fit(ratios_spanned, args.basekey, args.midi_truncate)
        print_step_hz_table(ratios_eff, basenote)

        # CSD: con --interval-zero considera l'ambitus (non ripetibile); altrimenti ignora ambitus e usa intervallo di ottava
        if getattr(args, "interval_zero", False):
            csd_input = ratios_spanned
            csd_interval = 0.0
        else:
            csd_input = ratios
            csd_interval = 2.0
        csd_ratios, csd_basekey = ensure_midi_fit(csd_input, args.basekey, args.midi_truncate)
        fnum, existed = write_cpstun_table(args.output_file, csd_ratios, csd_basekey, basenote, csd_interval)
        export_base = args.output_file if (not existed or fnum <= 1) else f"{args.output_file}_{fnum}"
        export_system_tables(export_base, ratios_eff, basekey_eff, basenote)
        try:
            diapason_hz = float(args.diapason)
        except TypeError:
            diapason_hz = 440.0
        export_comparison_tables(export_base, ratios_eff, basekey_eff, basenote, diapason_hz,
                                 compare_fund_hz=compare_fund_hz, tet_align=args.compare_tet_align)
        if args.export_tun:
            write_tun_file(export_base, ratios_eff, basekey_eff, basenote)
        return

    if args.danielou is not None:
        # Calcolo manuale: genera solo i rapporti per le terne specificate
        ratios: list[float] = []
        for (a, b, c) in args.danielou:
            ratios.extend(danielou_from_exponents(a, b, c, reduce_octave=(not args.no_reduce)))
        ratios_spanned = repeat_ratios(ratios, args.span, 2.0)
        ratios_eff, basekey_eff = ensure_midi_fit(ratios_spanned, args.basekey, args.midi_truncate)
        print_step_hz_table(ratios_eff, basenote)

        # CSD: con --interval-zero considera l'ambitus (non ripetibile); altrimenti ignora ambitus e usa intervallo di ottava
        if getattr(args, "interval_zero", False):
            csd_input = ratios_spanned
            csd_interval = 0.0
        else:
            csd_input = ratios
            csd_interval = 2.0
        csd_ratios, csd_basekey = ensure_midi_fit(csd_input, args.basekey, args.midi_truncate)
        fnum, existed = write_cpstun_table(args.output_file, csd_ratios, csd_basekey, basenote, csd_interval)
        export_base = args.output_file if (not existed or fnum <= 1) else f"{args.output_file}_{fnum}"
        export_system_tables(export_base, ratios_eff, basekey_eff, basenote)
        try:
            diapason_hz = float(args.diapason)
        except TypeError:
            diapason_hz = 440.0
        export_comparison_tables(export_base, ratios_eff, basekey_eff, basenote, diapason_hz,
                                 compare_fund_hz=compare_fund_hz, tet_align=args.compare_tet_align)
        if args.export_tun:
            write_tun_file(export_base, ratios_eff, basekey_eff, basenote)
        return

    # Sistema geometrico: se specificato, priorità rispetto a --et
    if args.geometric:
        gen_str, steps_str = args.geometric
        try:
            steps = int(steps_str)
            if steps <= 0:
                print("Numero di passi deve essere > 0 per --geometric.")
                return
        except ValueError:
            print("Numero di passi non valido per --geometric.")
            return
        # parsing generatore
        try:
            gen_val = int_or_fraction(gen_str)
        except argparse.ArgumentTypeError:
            try:
                gen_val = float(gen_str)
            except ValueError:
                print("Generatore non valido per --geometric.")
                return
        # normalizza tipo
        if isinstance(gen_val, int):
            gen_ratio = Fraction(gen_val, 1)
        elif isinstance(gen_val, Fraction):
            gen_ratio = gen_val
        else:
            gen_ratio = float(gen_val)

        # validazione generatore > 0
        try:
            gen_numeric = float(gen_ratio) if not isinstance(gen_ratio, Fraction) else float(gen_ratio)
        except TypeError:
            print("Generatore non numerico per --geometric.")
            return
        if gen_numeric <= 0:
            print("Il generatore deve essere > 0 per --geometric.")
            return

        ratios: list[float] = []
        for i in range(steps):
            # potenza del generatore
            if isinstance(gen_ratio, Fraction):
                r = gen_ratio ** i
            else:
                r = (float(gen_ratio)) ** i
            # riduzione all'ottava di default
            r_reduced = reduce_to_octave(r) if not args.no_reduce else r
            r_value = float(r_reduced) if isinstance(r_reduced, Fraction) else float(r_reduced)
            ratios.append(r_value)
        # Applica span su ottave
        ratios_spanned = repeat_ratios(ratios, args.span, 2.0)
        ratios_eff, basekey_eff = ensure_midi_fit(ratios_spanned, args.basekey, args.midi_truncate)
        print_step_hz_table(ratios_eff, basenote)

        # CSD: con --interval-zero considera l'ambitus (non ripetibile); altrimenti ignora ambitus e usa intervallo di ottava
        if getattr(args, "interval_zero", False):
            csd_input = ratios_spanned
            csd_interval = 0.0
        else:
            csd_input = ratios
            csd_interval = 2.0
        csd_ratios, csd_basekey = ensure_midi_fit(csd_input, args.basekey, args.midi_truncate)
        fnum, existed = write_cpstun_table(args.output_file, csd_ratios, csd_basekey, basenote, csd_interval)
        export_base = args.output_file if not existed else f"{args.output_file}_{fnum}"
        export_system_tables(export_base, ratios_eff, basekey_eff, basenote)
        try:
            diapason_hz = float(args.diapason)
        except TypeError:
            diapason_hz = 440.0
        export_comparison_tables(export_base, ratios_eff, basekey_eff, basenote, diapason_hz,
                                 compare_fund_hz=compare_fund_hz, tet_align=args.compare_tet_align)
        if args.export_tun:
            write_tun_file(export_base, ratios_eff, basekey_eff, basenote)
        return

    # Validazione presenza -et
    if not args.et:
        print("Temperamento equabile non specificato. Usa --et indice cents")
        return

    index, cents = args.et

    # Controlli di base
    if index <= 0 or (isinstance(cents, (int, float)) and cents <= 0):
        print("Indice della radice o valore di cents non valido.")
        return

    # Se l'utente ha passato una frazione, convertirla in cents
    if is_fraction_type(args.et[1]):
        print("rilevato numero razionale espresso in frazione, conversione in cents...")
        cents = fraction_to_cents(args.et[1])
        print(f"Conversione di {Fraction(args.et[1])} in cents: {cents}")

    print(f"index: {index}, cents: {cents}")

    try:
        ratio = ratio_et(index, cents)
        print(f"Ratio: {ratio}")
    except ZeroDivisionError:
        print("errore di divisione per zero")
        return

    base_ratios: list[float] = []
    for i in range(index):
        r_value = float(ratio ** i)
        base_ratios.append(r_value)
    # Intervallo di ripetizione per ET: exp(cents/ELLIS_CONVERSION_FACTOR)
    interval_factor = math.exp((cents / ELLIS_CONVERSION_FACTOR))
    # Span per stampa e confronti
    ratios_spanned = repeat_ratios(base_ratios, args.span, interval_factor)
    ratios_eff, basekey_eff = ensure_midi_fit(ratios_spanned, args.basekey, args.midi_truncate)
    print_step_hz_table(ratios_eff, basenote)

    # CSD: con --interval-zero considera l'ambitus (non ripetibile); altrimenti ignora ambitus e usa intervallo del sistema
    if getattr(args, "interval_zero", False):
        csd_input = ratios_spanned
        csd_interval = 0.0
    else:
        csd_input = base_ratios
        csd_interval = interval_factor
    csd_ratios, csd_basekey = ensure_midi_fit(csd_input, args.basekey, args.midi_truncate)
    fnum, existed = write_cpstun_table(args.output_file, csd_ratios, csd_basekey, basenote, csd_interval)

    export_base = args.output_file if not existed else f"{args.output_file}_{fnum}"
    export_system_tables(export_base, ratios_eff, basekey_eff, basenote)
    try:
        diapason_hz = float(args.diapason)
    except TypeError:
        diapason_hz = 440.0
    export_comparison_tables(export_base, ratios_eff, basekey_eff, basenote, diapason_hz,
                             compare_fund_hz=compare_fund_hz, tet_align=args.compare_tet_align)
    if args.export_tun:
        write_tun_file(export_base, ratios_eff, basekey_eff, basenote)
        
if __name__ == "__main__":
    main()