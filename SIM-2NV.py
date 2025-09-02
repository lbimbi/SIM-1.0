#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SIM - Sistemi intonazione musicale
Copyright (c) 2025 Luca Bimbi
Distribuito secondo la licenza MIT - vedi il file LICENSE per i dettagli

Nome programma: SIM - Sistemi intonazione musicale
Versione: 1.5
Autore: LUCA BIMBI
Data: 2025-08-30
"""

import argparse
import sys
import math
import re
import os
import shutil
from fractions import Fraction
from typing import List, Tuple, Optional, Union, Iterable, Any

# Metadati di modulo
__program_name__ = "SIM"
__version__ = "1.5"
__author__ = "LUCA BIMBI"
__date__ = "2025-08-30"

# Costanti
ELLIS_CONVERSION_FACTOR = 1200 / math.log(2)
RATIO_EPS = 1e-9
DEFAULT_DIAPASON = 440.0
DEFAULT_BASEKEY = 60
DEFAULT_OCTAVE = 2.0
MAX_HARMONIC_HZ = 10000.0
MIN_SUBHARMONIC_HZ = 16.0
PROXIMITY_THRESHOLD_HZ = 17.0
MIDI_MIN = 0
MIDI_MAX = 127
MIDI_A4 = 69
SEMITONES_PER_OCTAVE = 12

# Pattern per parsing file .csd
PATTERN = re.compile(r"\bf\s*(\d+)\b")

# Tipo per valori numerici (int, float o Fraction)
Numeric = Union[int, float, Fraction]


def is_fraction_type(value: Any) -> bool:
    """Verifica se value è una frazione."""
    return isinstance(value, Fraction)


def fraction_to_cents(ratio: Fraction) -> int:
    """Converte un rapporto razionale in cents."""
    if ratio.denominator == 0:
        raise ValueError("Denominatore zero nella frazione")
    decimal = math.log(ratio.numerator / ratio.denominator)
    return round(decimal * ELLIS_CONVERSION_FACTOR)


def cents_to_fraction(cents: float) -> Fraction:
    """Converte cents in rapporto razionale approssimato."""
    return Fraction(math.exp(cents / ELLIS_CONVERSION_FACTOR)).limit_denominator(10000)


def reduce_to_octave(value: Numeric) -> Numeric:
    """Riduce un rapporto nell'ambito di un'ottava [1, 2)."""
    if isinstance(value, Fraction):
        two = Fraction(2, 1)
        while value >= two:
            value /= two
        while value < 1:
            value *= two
    else:
        value = float(value)
        while value >= 2.0:
            value /= 2.0
        while value < 1.0:
            value *= 2.0
    return value


def reduce_to_interval(value: Numeric, interval: Numeric) -> Numeric:
    """Riduce un rapporto nell'intervallo [1, interval)."""
    try:
        interval_num = float(interval)
    except (TypeError, ValueError):
        return value

    if not math.isfinite(interval_num) or interval_num <= 1.0:
        return value

    if isinstance(value, Fraction) and isinstance(interval, Fraction):
        one = Fraction(1, 1)
        while value >= interval:
            value /= interval
        while value < one:
            value *= interval
    else:
        v = float(value)
        i = float(interval)
        while v >= i:
            v /= i
        while v < 1.0:
            v *= i
        return v

    return value


def normalize_ratios(ratios: Iterable[Numeric], reduce_octave: bool = True) -> List[float]:
    """
    Normalizza una lista di rapporti:
    - Opzionalmente riduce nell'ottava [1, 2)
    - Elimina duplicati con tolleranza
    - Ordina in modo crescente
    """
    processed = []
    seen = []

    for r in ratios:
        v = reduce_to_octave(r) if reduce_octave else r
        v_float = float(v)

        # Controlla duplicati con tolleranza
        if not any(abs(v_float - s) <= RATIO_EPS for s in seen):
            seen.append(v_float)
            processed.append(v_float)

    return sorted(processed)


def repeat_ratios(ratios: List[float], span: int, interval_factor: float) -> List[float]:
    """Ripete la serie di rapporti su un certo ambitus."""
    try:
        s = int(span)
    except (TypeError, ValueError):
        s = 1

    if s <= 1:
        return list(ratios)

    if not isinstance(interval_factor, (int, float)) or interval_factor <= 0:
        return list(ratios)

    out = []
    for k in range(s):
        factor_k = interval_factor ** k
        out.extend(float(r) * factor_k for r in ratios)

    return out


def pow_fraction(fr: Numeric, k: int) -> Numeric:
    """Eleva una Fraction/float ad un esponente intero."""
    k = int(k)
    return fr ** k if isinstance(fr, Fraction) else float(fr) ** k


def build_natural_ratios(a_max: int, b_max: int, reduce_octave: bool = True) -> List[float]:
    """Genera rapporti del sistema naturale (4:5:6)."""
    three_over_two = Fraction(3, 2)
    five_over_four = Fraction(5, 4)
    vals = []

    for a in range(-abs(a_max), abs(a_max) + 1):
        for b in range(-abs(b_max), abs(b_max) + 1):
            r = pow_fraction(three_over_two, a) * pow_fraction(five_over_four, b)
            vals.append(r)

    return normalize_ratios(vals, reduce_octave=reduce_octave)


def build_danielou_ratios(full_grid: bool = False, reduce_octave: bool = True) -> List[float]:
    """Genera rapporti del sistema Danielou."""
    six_over_five = Fraction(6, 5)
    three_over_two = Fraction(3, 2)
    vals = []

    if full_grid:
        # Mappatura basata sugli appunti di Danielou
        series_spec = {
            0: (5, 5),  # Prima serie su 1/1
            1: (3, 4),  # Seconda serie su 6/5
            -1: (4, 3),  # Terza serie su 5/3
            2: (4, 4),  # Quarta serie su 36/25
            -2: (4, 4),  # Quinta serie su 25/18
            3: (4, 0),  # Sesta serie su 216/125
            -3: (0, 4),  # Settima serie su 125/108
        }

        for a, (desc_n, asc_n) in series_spec.items():
            # Quinte discendenti
            for b in range(-desc_n, 0):
                r = pow_fraction(six_over_five, a) * pow_fraction(three_over_two, b)
                vals.append(r)
            # Centro serie
            vals.append(pow_fraction(six_over_five, a))
            # Quinte ascendenti
            for b in range(1, asc_n + 1):
                r = pow_fraction(six_over_five, a) * pow_fraction(three_over_two, b)
                vals.append(r)
    else:
        # Sottoinsieme dimostrativo
        vals.append(Fraction(1, 1))
        # Asse delle quinte
        for b in range(-5, 6):
            vals.append(pow_fraction(three_over_two, b))
        # Terze minori armoniche
        for k in range(1, 4):
            vals.append(pow_fraction(six_over_five, k))
        # Seste maggiori armoniche
        five_over_three = Fraction(5, 3)
        for k in range(1, 4):
            vals.append(pow_fraction(five_over_three, k))

    normalized = normalize_ratios(vals, reduce_octave=reduce_octave)

    if full_grid and reduce_octave:
        # Costruisci lista finale di 53 gradi
        base = list(normalized)
        if not base or abs(base[0] - 1.0) > RATIO_EPS:
            base.append(1.0)
            base = sorted(base)

        base_52_unique = []
        for v in base[:52]:
            if not base_52_unique or abs(v - base_52_unique[-1]) > RATIO_EPS:
                base_52_unique.append(v)

        # Estendi se necessario
        if len(base_52_unique) < 52 and len(base) > len(base_52_unique):
            for v in base[len(base_52_unique):]:
                if all(abs(v - u) > RATIO_EPS for u in base_52_unique):
                    base_52_unique.append(v)
                    if len(base_52_unique) >= 52:
                        break

        normalized = base_52_unique[:52] + [2.0]

    return normalized


def danielou_from_exponents(a: int, b: int, c: int, reduce_octave: bool = True) -> List[float]:
    """Calcola un singolo rapporto del sistema Danielou."""
    six_over_five = Fraction(6, 5)
    three_over_two = Fraction(3, 2)
    two = Fraction(2, 1)

    r = (pow_fraction(six_over_five, a) *
         pow_fraction(three_over_two, b) *
         pow_fraction(two, c))

    if reduce_octave:
        r = reduce_to_octave(r)

    return [float(r)]


def ratio_et(index: int, cents: Union[int, Fraction]) -> float:
    """Calcola il rapporto della radice per un temperamento equabile."""
    decimal_number = math.exp(cents / ELLIS_CONVERSION_FACTOR)
    return decimal_number ** (1 / index)


def int_or_fraction(value: str) -> Union[int, Fraction]:
    """Parser per interi o frazioni."""
    try:
        return int(value)
    except ValueError:
        try:
            return Fraction(value)
        except ValueError:
            raise argparse.ArgumentTypeError(f"'{value}' non è un intero o frazione valida.")


def parse_interval_value(value: str) -> Union[Fraction, float]:
    """Parsa un valore di intervallo per --geometric.

    Regola: un intero senza suffisso è interpretato come cents; frazioni e float sono rapporti.
    Esempi: 700 -> cents; 700c -> cents; 3/2 -> rapporto; 2.0 -> rapporto.
    """
    if value is None:
        return Fraction(2, 1)

    s = str(value).strip().lower()

    # Controlla suffissi cents
    cents_suffixes = ['cents', 'cent', 'c']
    for suffix in cents_suffixes:
        if s.endswith(suffix):
            num_str = s[:-len(suffix)].strip()
            try:
                cents_val = float(num_str)
            except ValueError:
                raise argparse.ArgumentTypeError(f"Valore cents non valido: '{value}'")
            ratio = cents_to_fraction(cents_val)
            if float(ratio) <= 1.0:
                raise argparse.ArgumentTypeError("L'intervallo in cents deve essere > 0")
            return ratio

    # Prova int o frazione
    try:
        val = int_or_fraction(s)
        if isinstance(val, int):
            # Intero puro => cents
            cents_val = float(val)
            ratio = cents_to_fraction(cents_val)
            if float(ratio) <= 1.0:
                raise argparse.ArgumentTypeError("L'intervallo in cents deve essere > 0")
        else:
            ratio = val
            if float(ratio) <= 1.0:
                raise argparse.ArgumentTypeError("L'intervallo (rapporto) deve essere > 1")
        return ratio
    except argparse.ArgumentTypeError:
        # Prova float come rapporto
        try:
            f = float(s)
            if not math.isfinite(f) or f <= 1.0:
                raise argparse.ArgumentTypeError("L'intervallo (rapporto) deve essere > 1")
            return f
        except ValueError:
            raise argparse.ArgumentTypeError(f"Intervallo non valido: '{value}'")


def parse_danielou_tuple(value: str) -> Tuple[int, int, int]:
    """Parser per terne Danielou 'a,b,c'."""
    if value is None:
        return (0, 0, 1)

    s = str(value).strip()

    # Rimuove parentesi esterne
    if s and s[0] in '[({' and s[-1] in '])}':
        s = s[1:-1].strip()

    # Normalizza separatori
    for sep in (':', ';'):
        s = s.replace(sep, ',')

    # Split e pulizia
    s = s.strip(',')
    parts = [p.strip() for p in s.split(',') if p.strip()]

    # Caso singolo intero
    if len(parts) == 1:
        try:
            return (int(parts[0]), 0, 1)
        except ValueError:
            raise argparse.ArgumentTypeError("Formato non valido per --danielou")

    if len(parts) != 3:
        raise argparse.ArgumentTypeError("Formato non valido per --danielou. Usa 'a,b,c'")

    try:
        return tuple(int(p) for p in parts)
    except ValueError:
        raise argparse.ArgumentTypeError("Esponenti non validi per --danielou")


def note_name_or_frequency(value: str) -> Union[float, str]:
    """Parser per nome nota o frequenza."""
    try:
        return float(value)
    except (ValueError, TypeError):
        return str(value)


def convert_note_name_to_midi(note_name: str) -> int:
    """Converte nome nota in valore MIDI."""
    note_map = {'C': 0, 'D': 2, 'E': 4, 'F': 5, 'G': 7, 'A': 9, 'B': 11}
    note_name = note_name.strip().upper()

    if not note_name:
        raise ValueError("Nome nota vuoto")

    note = note_name[0]
    if note not in note_map:
        raise ValueError(f"Nome nota non valido: {note_name}")

    alteration = 0
    index_octave = 1

    if len(note_name) > 1:
        if note_name[1] == '#':
            alteration = 1
            index_octave = 2
        elif note_name[1] == 'B':
            alteration = -1
            index_octave = 2

    try:
        octave = int(note_name[index_octave:])
    except (ValueError, IndexError):
        raise ValueError(f"Formato ottava non valido in: {note_name}")

    midi_value = (octave + 1) * SEMITONES_PER_OCTAVE + note_map[note] + alteration

    if not (MIDI_MIN <= midi_value <= MIDI_MAX):
        raise ValueError(f"Nota MIDI fuori range: {midi_value}")

    return midi_value


def convert_midi_to_hz(midi_value: int, diapason_hz: float = DEFAULT_DIAPASON) -> float:
    """Converte valore MIDI in frequenza Hz."""
    return diapason_hz * (2 ** ((float(midi_value) - MIDI_A4) / SEMITONES_PER_OCTAVE))


def file_exists(file_path: str) -> bool:
    """Verifica esistenza file."""
    exists = os.path.exists(file_path)
    print(f"File {'esistente' if exists else 'non esistente'}: {file_path}")
    return exists


def parse_file(file_name: str) -> int:
    """Trova il massimo numero di tabella 'f N' nel file."""
    max_num = 0
    try:
        with open(file_name, "r") as file:
            for line in file:
                for match in PATTERN.finditer(line):
                    num = int(match.group(1))
                    max_num = max(max_num, num)
    except FileNotFoundError:
        print(f"Errore: file non trovato: {file_name}")
    return max_num


def ensure_midi_fit(ratios: List[float], basekey: int,
                    prefer_truncate: bool) -> Tuple[List[float], int]:
    """Assicura che i rapporti rientrino nel range MIDI."""
    n = len(ratios)

    try:
        bk = int(basekey)
    except (TypeError, ValueError):
        bk = 0

    if n > 128:
        if not prefer_truncate:
            print(f"WARNING: numero di passi ({n}) eccede 128. Verranno mantenuti solo i primi 128.")
            prefer_truncate = True
        n = 128
        ratios = ratios[:n]

    if prefer_truncate:
        bk = max(MIDI_MIN, min(MIDI_MAX, bk))
        max_len = 128 - bk
        if len(ratios) > max_len:
            print(f"WARNING: serie eccede il limite MIDI. Troncati {len(ratios) - max_len} passi.")
            ratios = ratios[:max_len]
        return ratios, bk
    else:
        allowed_min = MIDI_MIN
        allowed_max = MIDI_MAX - (n - 1)
        if allowed_max < allowed_min:
            allowed_max = allowed_min

        bk_eff = bk
        if bk < allowed_min or bk > allowed_max:
            bk_eff = max(allowed_min, min(allowed_max, bk))
            print(f"WARNING: basekey adattata da {bk} a {bk_eff} per includere tutti i passi.")

        return ratios, bk_eff


def write_cpstun_table(output_base: str, ratios: List[float], basekey: int,
                       basefrequency: float, interval_value: Optional[float] = None) -> Tuple[int, bool]:
    """Crea o appende una tabella cpstun in un file Csound .csd."""
    csd_path = f"{output_base}.csd"
    skeleton = (
        "<CsoundSynthesizer>\n"
        "<CsOptions>\n\n</CsOptions>\n"
        "<CsInstruments>\n\n</CsInstruments>\n"
        "<CsScore>\n\n</CsScore>\n"
        "</CsoundSynthesizer>\n"
    )

    existed_before = file_exists(csd_path)

    if not existed_before:
        try:
            with open(csd_path, "w") as f:
                f.write(skeleton)
        except IOError as e:
            print(f"Errore creazione file CSD: {e}")
            return 0, existed_before

    try:
        with open(csd_path, "r", encoding="utf-8") as f:
            content = f.read()
    except IOError as e:
        print(f"Errore lettura CSD: {e}")
        return 0, existed_before

    fnum = parse_file(csd_path) + 1

    # Ordina i rapporti
    try:
        ratios_sorted = sorted(float(r) for r in ratios)
    except (TypeError, ValueError):
        ratios_sorted = [float(r) for r in ratios]

    # Determina parametri cpstun
    numgrades = len(ratios_sorted)

    if interval_value is not None and isinstance(interval_value, (int, float)):
        interval = 0.0 if float(interval_value) <= 0 else float(interval_value)
    else:
        try:
            rmin = min(ratios_sorted) if ratios_sorted else 1.0
            rmax = max(ratios_sorted) if ratios_sorted else 1.0
            interval = 2.0 if (rmin >= 1.0 - RATIO_EPS and rmax <= 2.0 + RATIO_EPS) else 0.0
        except (TypeError, ValueError):
            interval = 0.0

    # Costruisci lista dati
    data_list = [
        str(numgrades),
        f"{float(interval):.10g}",
        f"{float(basefrequency):.10g}",
        str(int(basekey))
    ]
    data_list.extend(f"{r:.10f}" for r in ratios_sorted)

    size = len(data_list)

    # Costruisci righe
    prefix = f"f {fnum} 0 {size} -2 "
    positions = []
    col = len(prefix)
    for t in data_list:
        positions.append(col)
        col += len(t) + 1

    def build_aligned_comment(label_map: List[Tuple[int, str]]) -> str:
        line = ";"
        base_offset = 1
        for idx, text in label_map:
            target = base_offset + (positions[idx] if 0 <= idx < len(positions) else len(prefix))
            if target < len(line):
                target = len(line) + 1
            line += " " * (target - len(line)) + text
        return line + "\n"

    header_comment = (
            build_aligned_comment([(0, "numgrades"), (2, "basefreq"), (4, "rapporti-di-intonazione .......")]) +
            build_aligned_comment([(1, "interval"), (3, "basekey")]) +
            f"; tabella cpstun generata | basekey={basekey} basefrequency={basefrequency:.6f}Hz\n"
    )
    f_line = prefix + " ".join(data_list) + "\n"

    # Inserisci prima di </CsScore>
    insert_marker = "</CsScore>"
    idx = content.rfind(insert_marker)
    if idx == -1:
        content += f"\n<CsScore>\n{header_comment}{f_line}</CsScore>\n"
    else:
        content = content[:idx] + header_comment + f_line + content[idx:]

    try:
        with open(csd_path, "w", encoding="utf-8") as f:
            f.write(content)
        print(f"Tabella cpstun (f {fnum}) salvata in {csd_path}")
    except IOError as e:
        print(f"Errore scrittura CSD: {e}")

    return fnum, existed_before


def write_tun_file(output_base: str, ratios: List[float], basekey: int,
                   basefrequency: float) -> None:
    """Esporta un file .tun (AnaMark TUN) con valori espressi in cents assoluti riferiti a 8.1757989156437073336 Hz.
    Struttura: [Tuning] + 128 righe "note X=Y" (cents assoluti)."""
    tun_path = f"{output_base}.tun"
    lines = [
        "[Tuning]",
        # f"; basekey={basekey} basefrequency={basefrequency:.6f}Hz | valori in cents assoluti riferiti a 8.1757989156437073336 Hz"
    ]

    def tet_freq(offset_semitones: int) -> float:
        return basefrequency * (2.0 ** (offset_semitones / 12.0))

    # Ordina i rapporti per garantire valori crescenti nel segmento custom
    try:
        ratios_sorted = sorted(float(r) for r in ratios)
    except (TypeError, ValueError):
        ratios_sorted = [float(r) for r in ratios]

    def note_freq(n: int) -> float:
        if 0 <= basekey <= 127 and basekey <= n < basekey + len(ratios_sorted):
            idx = n - basekey
            if 0 <= idx < len(ratios_sorted):
                return basefrequency * float(ratios_sorted[idx])
            else:
                return tet_freq(n - basekey)
        else:
            return tet_freq(n - basekey)

    # Riferimento assoluto AnaMark
    f_ref = 8.1757989156437073336

    for n in range(128):
        f = note_freq(n)
        if isinstance(f, (int, float)) and f > 0:
            cents = 1200.0 * math.log2(f / f_ref)
        else:
            cents = 0.0 if n == 0 else 0.0
        lines.append(f"note {n}={cents:.10f}")

    try:
        with open(tun_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")
        print(f"File .tun salvato in {tun_path}")
    except IOError as e:
        print(f"Errore scrittura .tun: {e}")


def export_system_tables(output_base: str, ratios: List[float], basekey: int,
                         basenote_hz: float) -> None:
    """Esporta tabelle del sistema generato."""
    # Calcola e ordina
    computed = [(basenote_hz * float(r), i, basekey + i, float(r))
                for i, r in enumerate(ratios)]
    computed.sort(key=lambda t: t[0])

    # Export testo
    txt_path = f"{output_base}_system.txt"
    try:
        headers = ["Step", "MIDI", "Ratio", "Hz"]
        rows = [[str(i), str(basekey + i), f"{r:.10f}", f"{hz:.6f}"]
                for i, (hz, _, _, r) in enumerate(computed)]

        # Calcola larghezze colonne
        widths = [len(h) for h in headers]
        for row in rows:
            for c, val in enumerate(row):
                widths[c] = max(widths[c], len(val))

        def fmt(vals: List[str]) -> str:
            return "  ".join(val.ljust(widths[i]) for i, val in enumerate(vals))

        with open(txt_path, "w", encoding="utf-8") as f:
            f.write(fmt(headers) + "\n")
            for row in rows:
                f.write(fmt(row) + "\n")
        print(f"Esportato: {txt_path}")
    except IOError as e:
        print(f"Errore scrittura {txt_path}: {e}")

    # Export Excel (opzionale)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "System"
        ws.append(headers)

        header_fill = PatternFill(start_color="FFDDDDDD", end_color="FFDDDDDD",
                                  fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        for i, (hz, _, _, r) in enumerate(computed):
            ws.append([i, basekey + i, r, hz])

        xlsx_path = f"{output_base}_system.xlsx"
        wb.save(xlsx_path)
        print(f"Esportato: {xlsx_path}")
    except ImportError:
        print("openpyxl non installato: export Excel saltato")
    except Exception as e:
        print(f"Errore export Excel: {e}")


def export_comparison_tables(output_base: str, ratios: List[float], basekey: int,
                             basenote_hz: float, diapason_hz: float,
                             compare_fund_hz: Optional[float] = None,
                             tet_align: str = "same",
                             subharm_fund_hz: Optional[float] = None) -> None:
    """Esporta tabelle di confronto con 12TET, serie armonica e subarmonica."""

    def freq_to_note_name(freq: float, a4_hz: float) -> str:
        """Converte frequenza in nome nota."""
        try:
            if not (freq > 0 and a4_hz > 0):
                return ""
            midi = int(round(MIDI_A4 + SEMITONES_PER_OCTAVE * math.log2(freq / a4_hz)))
        except (ValueError, OverflowError):
            return ""

        midi = max(MIDI_MIN, min(MIDI_MAX, midi))
        names = ["C", "C#", "D", "D#", "E", "F", "F#", "G", "G#", "A", "A#", "B"]
        name = names[midi % SEMITONES_PER_OCTAVE]
        octave = (midi // SEMITONES_PER_OCTAVE) - 1
        return f"{name}{octave}"

    base_cmp = compare_fund_hz if compare_fund_hz is not None else basenote_hz
    sub_base = subharm_fund_hz if subharm_fund_hz is not None else diapason_hz

    # Prepara dati ordinati
    computed = [(basenote_hz * float(r), i, basekey + i, float(r))
                for i, r in enumerate(ratios)]
    computed.sort(key=lambda t: t[0])

    custom_list = computed
    n_rows = len(custom_list)

    # Serie armonica
    harm_vals = []
    n = 1
    while True:
        val = base_cmp * n
        if val > MAX_HARMONIC_HZ:
            break
        harm_vals.append(val)
        n += 1

    # Serie subarmonica
    sub_desc = []
    m = 1
    while True:
        val = sub_base / m
        if val < MIN_SUBHARMONIC_HZ:
            break
        sub_desc.append(val)
        m += 1
    sub_vals = list(reversed(sub_desc))

    # Filtro subarmoniche
    min_custom = custom_list[0][0] if n_rows > 0 else base_cmp
    cutoff_sub = max(MIN_SUBHARMONIC_HZ, min_custom)
    sub_vals = [v for v in sub_vals if v >= cutoff_sub]

    # Allineamento sequenze
    def align_sequence(seq: List[float], customs: List[float]) -> List[Optional[float]]:
        out = [None] * len(customs)
        p = 0
        for i in range(len(customs)):
            low = customs[i]
            high = customs[i + 1] if i + 1 < len(customs) else float('inf')
            if p < len(seq) and low <= seq[p] < high:
                out[i] = seq[p]
                p += 1
        return out

    custom_hz_list = [c[0] for c in custom_list]
    harm_aligned = align_sequence(harm_vals, custom_hz_list)
    sub_aligned = align_sequence(sub_vals, custom_hz_list)

    # Export testo
    txt_path = f"{output_base}_compare.txt"
    try:
        headers = [
            "Step", "MIDI", "Ratio",
            "Custom_Hz", "Harmonic_Hz", "DeltaHz_Harm",
            "Subharm_Hz", "DeltaHz_Sub",
            "TET_Hz", "TET_Note", "DeltaHz_TET"
        ]

        rows = []
        for row_i, (custom_hz, i, midi, r) in enumerate(custom_list):
            harm_val = harm_aligned[row_i]
            sub_val = sub_aligned[row_i]

            # 12TET
            if custom_hz > 0 and base_cmp > 0:
                n = round(SEMITONES_PER_OCTAVE * math.log2(custom_hz / base_cmp))
                tet_val = base_cmp * (2.0 ** (n / SEMITONES_PER_OCTAVE))
            else:
                tet_val = base_cmp

            # Formatta valori
            harm_str = f"{harm_val:.6f}" if harm_val is not None else ""
            d_har_str = f"{(custom_hz - harm_val):.6f}" if harm_val is not None else ""
            approx = "≈" if harm_val is not None and abs(custom_hz - harm_val) < PROXIMITY_THRESHOLD_HZ else ""

            sub_str = f"{sub_val:.6f}" if sub_val is not None else ""
            d_sub_str = f"{(custom_hz - sub_val):.6f}" if sub_val is not None else ""

            tet_str = f"{tet_val:.6f}"
            tet_note = freq_to_note_name(tet_val, diapason_hz)
            d_tet_str = f"{(custom_hz - tet_val):.6f}"

            rows.append([
                str(row_i), str(basekey + row_i), f"{r:.10f}",
                f"{custom_hz:.6f}{approx}",
                f"{harm_str}{approx if harm_str else ''}",
                d_har_str, sub_str, d_sub_str,
                tet_str, tet_note, d_tet_str
            ])

        # Calcola larghezze
        widths = [len(h) for h in headers]
        for row in rows:
            for c, val in enumerate(row):
                widths[c] = max(widths[c], len(val))

        def fmt(vals: List[str]) -> str:
            return "  ".join(val.ljust(widths[i]) for i, val in enumerate(vals))

        with open(txt_path, "w", encoding="utf-8") as f:
            f.write(fmt(headers) + "\n")
            for row in rows:
                f.write(fmt(row) + "\n")
        print(f"Esportato: {txt_path}")
    except IOError as e:
        print(f"Errore scrittura {txt_path}: {e}")

    # Export Excel (opzionale)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Compare"

        headers_xl = [
            "Step", "MIDI", "Ratio",
            "Custom_Hz", "Harmonic_Hz", "|DeltaHz_Harm|",
            "Subharm_Hz", "|DeltaHz_Sub|",
            "TET_Hz", "TET_Note", "|DeltaHz_TET|"
        ]
        ws.append(headers_xl)

        header_fill = PatternFill(start_color="FFCCE5FF", end_color="FFCCE5FF",
                                  fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        # Colori serie
        fill_custom = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC",
                                  fill_type="solid")
        fill_harm = PatternFill(start_color="FFCCFFCC", end_color="FFCCFFCC",
                                fill_type="solid")
        fill_sub = PatternFill(start_color="FFFFFFCC", end_color="FFFFFFCC",
                               fill_type="solid")
        fill_tet = PatternFill(start_color="FFCCE5FF", end_color="FFCCE5FF",
                               fill_type="solid")

        for row_i, (custom_hz, i, midi, r) in enumerate(custom_list):
            harm_val = harm_aligned[row_i]
            sub_val = sub_aligned[row_i]

            if custom_hz > 0 and base_cmp > 0:
                n = round(SEMITONES_PER_OCTAVE * math.log2(custom_hz / base_cmp))
                tet_val = base_cmp * (2.0 ** (n / SEMITONES_PER_OCTAVE))
            else:
                tet_val = base_cmp

            harm_cell = harm_val if harm_val is not None else None
            d_har_cell = abs(custom_hz - harm_val) if harm_val is not None else None

            sub_cell = sub_val if sub_val is not None else None
            d_sub_cell = abs(custom_hz - sub_val) if sub_val is not None else None

            tet_note = freq_to_note_name(tet_val, diapason_hz)
            d_tet_cell = abs(custom_hz - tet_val)

            ws.append([row_i, basekey + row_i, r, custom_hz, harm_cell, d_har_cell,
                       sub_cell, d_sub_cell, tet_val, tet_note, d_tet_cell])

            row = ws.max_row
            ws.cell(row=row, column=4).fill = fill_custom
            ws.cell(row=row, column=5).fill = fill_harm
            ws.cell(row=row, column=7).fill = fill_sub
            ws.cell(row=row, column=9).fill = fill_tet
            ws.cell(row=row, column=10).fill = fill_tet

        xlsx_path = f"{output_base}_compare.xlsx"
        wb.save(xlsx_path)
        print(f"Esportato: {xlsx_path}")
    except ImportError:
        print("openpyxl non installato: export Excel saltato")
    except Exception as e:
        print(f"Errore export Excel: {e}")


def print_step_hz_table(ratios: List[float], basenote_hz: float) -> None:
    """Stampa tabella multi-colonna con Step/Hz."""
    pairs = [(str(i + 1), f"{basenote_hz * float(r):.2f}")
             for i, r in enumerate(ratios)]

    if not pairs:
        print("Step Hz")
        return

    # Calcola larghezze
    step_w = max(len("Step"), max(len(p[0]) for p in pairs))
    hz_w = max(len("Hz"), max(len(p[1]) for p in pairs))
    cell_w = step_w + 1 + hz_w
    gap = 3

    # Larghezza terminale
    try:
        term_w = shutil.get_terminal_size(fallback=(80, 24)).columns
    except (AttributeError, ValueError):
        term_w = 80

    # Calcola colonne
    cols = max(1, (term_w + gap) // (cell_w + gap)) if cell_w < term_w else 1
    rows_count = math.ceil(len(pairs) / cols)

    # Header
    header_cell = "Step".ljust(step_w) + " " + "Hz".ljust(hz_w)
    print((" " * gap).join([header_cell] * cols))

    # Righe
    for r in range(rows_count):
        line_cells = []
        for c in range(cols):
            idx = c * rows_count + r
            if idx < len(pairs):
                s, h = pairs[idx]
                cell = s.rjust(step_w) + " " + h.rjust(hz_w)
                line_cells.append(cell)
        if line_cells:
            print((" " * gap).join(line_cells))


def process_tuning_system(args: argparse.Namespace, basenote: float) -> Optional[Tuple[List[float], float]]:
    """Processa il sistema di intonazione e ritorna (ratios, interval) o None."""

    # Sistema naturale
    if args.natural:
        try:
            a_max = int(args.natural[0])
            b_max = int(args.natural[1])
            if a_max < 0 or b_max < 0:
                print("A_MAX e B_MAX devono essere >= 0")
                return None
            ratios = build_natural_ratios(a_max, b_max, not args.no_reduce)
            return ratios, DEFAULT_OCTAVE
        except (TypeError, ValueError):
            print("Valori non validi per --natural")
            return None

    # Sistema Danielou
    if args.danielou_all:
        ratios = build_danielou_ratios(True, not args.no_reduce)
        return ratios, DEFAULT_OCTAVE

    if args.danielou is not None:
        ratios = []
        for (a, b, c) in args.danielou:
            ratios.extend(danielou_from_exponents(a, b, c, not args.no_reduce))
        return ratios, DEFAULT_OCTAVE

    # Sistema geometrico
    if args.geometric:
        parts = list(args.geometric)
        if len(parts) != 3:
            print("Uso: --geometric GEN STEPS INTERVAL")
            return None

        try:
            interval_ratio = parse_interval_value(parts[2])
            steps = int(parts[1])
            if steps <= 0:
                print("Numero di passi deve essere > 0")
                return None

            gen_val = int_or_fraction(parts[0])
            gen_ratio = Fraction(gen_val, 1) if isinstance(gen_val, int) else gen_val
        except (argparse.ArgumentTypeError, ValueError) as e:
            try:
                gen_ratio = float(parts[0])
            except ValueError:
                print(f"Errore parsing geometrico: {e}")
                return None

        if float(gen_ratio) <= 0:
            print("Il generatore deve essere > 0")
            return None

        ratios = []
        for i in range(steps):
            r = pow_fraction(gen_ratio, i) if isinstance(gen_ratio, Fraction) else float(gen_ratio) ** i
            if not args.no_reduce:
                r = reduce_to_interval(r, interval_ratio)
            ratios.append(float(r))

        return ratios, float(interval_ratio)

    # Temperamento equabile (default)
    if args.et:
        index, cents = args.et
        if index <= 0 or (isinstance(cents, (int, float)) and cents <= 0):
            print("Indice o cents non valido")
            return None

        if is_fraction_type(cents):
            print("Conversione frazione in cents...")
            cents = fraction_to_cents(cents)
            print(f"Conversione di {Fraction(args.et[1])} in cents: {cents}")

        print(f"index: {index}, cents: {cents}")

        try:
            ratio = ratio_et(index, cents)
            print(f"Ratio: {ratio}")
        except ZeroDivisionError:
            print("Errore divisione per zero")
            return None

        ratios = [float(ratio ** i) for i in range(index)]
        interval_factor = math.exp(cents / ELLIS_CONVERSION_FACTOR)
        return ratios, interval_factor

    return None


def main():
    """Punto di ingresso principale."""
    parser = argparse.ArgumentParser(
        description=f"SIM. Assistente per sistemi di intonazione musicale. Versione {__version__}",
        epilog="Copyright (C) 2025 Luca Bimbi. Licenza MIT."
    )

    # Argomenti base
    parser.add_argument("-v", "--version", action="version",
                        version=f"%(prog)s {__version__}")
    parser.add_argument("--diapason", type=float, default=DEFAULT_DIAPASON,
                        help=f"Diapason in Hz (default: {DEFAULT_DIAPASON})")
    parser.add_argument("--basekey", type=int, default=DEFAULT_BASEKEY,
                        help=f"Nota base MIDI (default: {DEFAULT_BASEKEY})")
    parser.add_argument("--basenote", type=note_name_or_frequency, default="C4",
                        help="Nota di riferimento o frequenza in Hz")

    # Sistemi di intonazione
    parser.add_argument("--et", nargs=2, type=int_or_fraction, default=(12, 1200),
                        metavar=("INDEX", "INTERVAL"),
                        help="Temperamento equabile")
    parser.add_argument("--geometric", nargs=3, metavar=("GEN", "STEPS", "INTERVAL"),
                        help=(
                            "Sistema geometrico: INTERVAL intero=cents (es. 700); per rapporto usare 2/1 o 2.0; "
                            "accetta anche suffisso 'c' (es. 200c)."
                        ))
    parser.add_argument("--natural", nargs=2, type=int, metavar=("A_MAX", "B_MAX"),
                        help="Sistema naturale 4:5:6")
    parser.add_argument("--danielou", action="append", type=parse_danielou_tuple,
                        default=None, help="Sistema Danielou manuale")
    parser.add_argument("--danielou-all", action="store_true",
                        help="Griglia completa Danielou")

    # Opzioni
    parser.add_argument("--no-reduce", action="store_true",
                        help="Non ridurre all'ottava")
    parser.add_argument("--span", "--ambitus", dest="span", type=int, default=1,
                        help="Ripetizioni intervallo")
    parser.add_argument("--interval-zero", action="store_true",
                        help="Imposta interval=0 in cpstun")
    parser.add_argument("--export-tun", action="store_true",
                        help="Esporta file .tun")

    # Confronto
    parser.add_argument("--compare-fund", nargs="?", type=note_name_or_frequency,
                        default="basenote", const="basenote",
                        help="Fondamentale per confronto")
    parser.add_argument("--compare-tet-align", choices=["same", "nearest"],
                        default="same", help="Allineamento 12TET")
    parser.add_argument("--subharm-fund", type=note_name_or_frequency,
                        default="A5", help="Fondamentale subarmoniche")
    parser.add_argument("--midi-truncate", action="store_true",
                        help="Forza troncamento MIDI")

    # Output
    parser.add_argument("output_file", nargs="?", default=None,
                        help="File di output (default: out)")

    # Mostra help se nessun argomento
    if len(sys.argv) == 1:
        parser.print_help()
        return

    args = parser.parse_args()

    # Validazione base
    if isinstance(args.basenote, (int, float)) and args.basenote < 0:
        print("Nota di riferimento non valida")
        return

    if args.span is None or args.span < 1:
        args.span = 1

    # Calcola frequenza base
    if isinstance(args.basenote, float):
        basenote = args.basenote
    else:
        try:
            basenote_midi = convert_note_name_to_midi(args.basenote)
            basenote = convert_midi_to_hz(basenote_midi, args.diapason)
        except ValueError as e:
            print(f"Errore conversione nota: {e}")
            return

    # Fondamentali per confronto
    def parse_fund_hz(value, default):
        if isinstance(value, str) and value.lower() == "basenote":
            return basenote
        elif isinstance(value, float):
            return value
        else:
            try:
                midi = convert_note_name_to_midi(str(value))
                return convert_midi_to_hz(midi, args.diapason)
            except ValueError:
                return default

    compare_fund_hz = parse_fund_hz(args.compare_fund, basenote)
    subharm_fund_hz = parse_fund_hz(args.subharm_fund, args.diapason)


    # Processa sistema di intonazione
    result = process_tuning_system(args, basenote)
    if result is None:
        print("Nessun sistema di intonazione valido specificato")
        return

    ratios, interval = result

    # Applica span
    ratios_spanned = repeat_ratios(ratios, args.span, interval)
    ratios_eff, basekey_eff = ensure_midi_fit(ratios_spanned, args.basekey,
                                              args.midi_truncate)

    # Stampa tabella
    print_step_hz_table(sorted(ratios_eff), basenote)

    # Prepara dati per CSD
    if args.interval_zero:
        csd_input = ratios_spanned
        csd_interval = 0.0
    else:
        csd_input = ratios
        csd_interval = interval

    csd_ratios, csd_basekey = ensure_midi_fit(csd_input, args.basekey,
                                              args.midi_truncate)
    output_base = args.output_file or "out"
    fnum, existed = write_cpstun_table(output_base, csd_ratios,
                                       csd_basekey, basenote, csd_interval)

    # Export
    export_base = (output_base if not existed
                   else f"{output_base}_{fnum}")

    export_system_tables(export_base, ratios_eff, basekey_eff, basenote)
    export_comparison_tables(export_base, ratios_eff, basekey_eff, basenote,
                             args.diapason, compare_fund_hz,
                             args.compare_tet_align, subharm_fund_hz)

    if args.export_tun:
        write_tun_file(export_base, ratios_eff, basekey_eff, basenote)


if __name__ == "__main__":
    main()